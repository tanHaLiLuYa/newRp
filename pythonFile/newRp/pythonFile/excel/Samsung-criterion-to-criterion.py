
# encoding: utf-8
import PIL
import sys,os,openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
excel_path=r"E:\work\samsung\W47周\W47 MS AuditChecklist-20191115 - 副本.xlsx"
excel_new =r"E:\work\samsung\W47周\W47 MS AuditChecklist-20191115 - 副本2gx.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheets = wb.sheetnames
dic_item_value={}#字典 >列表 >字典
dic_item_criterion={}#字典 >列表 >字典
dic_item ={}#字典 >列表
list_item =[]#所有项目的合并列表
for sheetname in sheets:
    ws = wb[sheetname]
    max_row =ws.max_row
    list_item0=[]
    list_item1=[]
    list_item2=[]
    for row in range(3,max_row):
        item_value    = str(ws[f"B{row}"].value)#项目列》》》》》》B列
        list_value0  = item_value[:3]
        list_value1 = ws[f"C{row}"].value#分数值列》》》》》》》》C列
        list_value2 = ws[f"D{row}"].value#评分标准》》》》》》》》D列
        if "." in list_value0 and list_value1:
            list_item.append(item_value[3:])#加入所有的项目
            list_item0.append({item_value[3:]:list_value1})#加入项目 和 分数值 的字典
            list_item2.append({item_value[3:]:list_value2})#加入项目 和 评分标准 的字典
            list_item1.append(item_value[3:])#加入所有的项目 > 字典中的列表
    dic_item_value    [sheetname]=list_item0
    dic_item          [sheetname]=list_item1
    dic_item_criterion[sheetname]=list_item2


list_sole_item=[]#不重复的item列表
mid_list=[]
#从每个类型门店的item中一项一项的加入到list_sole_item
for n in range(0,len(list_item)+1):
    for s in sheets:
        if n <len(dic_item[s]):
            mid_list.append(dic_item[s][n])
    # [list_sole_item.append(i) for i in mid_list if not i in list_sole_item ]
    for i in mid_list:
        if i not in list_sole_item:
            list_sole_item.append(i)
# print(list_sole_item,len(list_sole_item))


#新建sheet  
ws1=wb.create_sheet("endsheet",0)
r=3
for k0,v0 in dic_item_value.items():
    #写入门店类型
    ft1 = Font(color=colors.BLACK,name="微软雅黑")
    ft  = Font(color=colors.RED,name="微软雅黑")
    ws1[f"B{r}"].value=k0
    if (r%2) == 0:  
        ws1[f"B{r}"].font =ft
    else:
        ws1[f"B{r}"].font =ft1
    #写入 item 和 value 
    for i in range(len(list_sole_item)):
        ws1[f"{get_column_letter(i+3)}2"].value =list_sole_item[i] 
        ws1[f"{get_column_letter(i+3)}2"].font =ft1
        for v1 in dic_item_value[k0]:
            for k2,v2 in v1.items():
                if k2 == list_sole_item[i]:
                    ws1[f"{get_column_letter(i+3)}{r}"].value = v2
                    if (r%2) == 0:
                        ws1[f"{get_column_letter(i+3)}{r}"].font = ft
                    else:
                        ws1[f"{get_column_letter(i+3)}{r}"].font = ft1
        # 写入criteria
        for v3 in dic_item_criterion[k0]:
            for k4,v4 in v3.items():
                if k4 == list_sole_item[i]:
                    ws1[f"{get_column_letter(i+3)}{7}"].value = v4
                    ws1[f"{get_column_letter(i+3)}{7}"].font = ft1

    ws1[f"A{r}"]=f"=sum(C{r}:{get_column_letter(len(list_sole_item)+2)}{r})"
    r = r +1


wb.save(excel_new)
if __name__ == '__main__':
    pass
