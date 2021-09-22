from openpyxl.utils.dataframe import dataframe_to_rows

import openpyxl
import os
from openpyxl.utils import get_column_letter, column_index_from_string

RootPath = r"C:\Users\16930\Desktop\python 测试"
pptx_path = os.path.join(
    RootPath, "Y'19 10月_MS Audit Monthly Report_091023_1620.pptx")

exl_path = os.path.join(RootPath, "W38-W41_过渡data_191028_1430.xlsx")

# print(new_pptx_path)
wb = openpyxl.load_workbook(exl_path, data_only=True, read_only=True)
ws = wb["Appendix得分和问题分析"]
area_ws = ws["C115:G126"]
# print(area_ws[0][0].value)
# for row in range(3,6):
#     for col in range(4,5):
#         print(area_ws[row][col].number_format)
