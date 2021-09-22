import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import PIL
import sys
from openpyxl.drawing.image import Image
import xlwings as xw

root_path = r"C:\Users\16930\Desktop\新建文件夹\照片"
middle_path = r"C:\Users\16930\Desktop\新建文件夹\照片\新建文件夹"
aim_path = r"C:\Users\16930\Desktop\新建文件夹\照片\aim"#里面的文件必须清空
file_name = r"最新导购员照片表.xlsx"
split_by = "CBI负责人"
#照片路径
pic_dir =r"G:\导购员20190828"
pic_dirs = os.walk(pic_dir)
excel_dirs =os.walk(middle_path)

file_path = os.path.join(root_path, file_name)
save_path = os.path.join(middle_path , file_name)


# 查找分列依据在哪个列的函数
def get_cell_by_content(worksheet, content):
    for row in range(1, 5):
        for col in range(1, 50):
            if worksheet.cell(row, col).value == content:
                if worksheet.cell(row, col).value != "":
                    return [row, col]
                else:
                    return [row+1, col]


# 打开原始表
o_wb = openpyxl.load_workbook(file_path, data_only=True)

ws_active = o_wb.active
content_cell = get_cell_by_content(ws_active, split_by)
print(content_cell)
max_row = ws_active.max_row
col_name = get_column_letter(content_cell[1])
row_min = content_cell[0]+1

# 获取第一个sheet的分表依据的列表
split_list = []
for cell in ws_active[f"{col_name}{row_min}:{col_name}{max_row}"]:
    if cell[0].value != None:
        split_list.append(cell[0].value)

split_list = list(set(split_list))
print(split_list)

# 遍历要分表的依据，然后把原始文件另存为一个一个新文件
for l in split_list:
    wb_new_name = save_path.replace(".xlsx", f"_{l}.xlsx")  # 新文件名

    print(wb_new_name)
    o_wb.save(wb_new_name)
    wb = openpyxl.load_workbook(wb_new_name)
    sheets = wb.sheetnames
    # 遍历新文件的各个sheet：
    for sheet in sheets:

        cell_position = get_cell_by_content(
            wb[sheet], split_by)  # 这是字符串所在的那个格的位置
        col_name = get_column_letter(cell_position[1])  # 这是那个格的列号

        print(f"在《{sheet}》里，「{split_by}」所在的位置为{cell_position}")
        row_max = wb[sheet].max_row  # 这是最大行号
        # 这是总部编码下面的那个行
        row_min = cell_position[0]

        print(f"{split_by}下面有字的第一行是{row_min}")
        

        for row_i in range(row_max, row_min, -1):
            v = wb[sheet][f"{col_name}{row_i}"].value
            if v == l:
                pass
            else:
                # print("将要删除这一行")
                
                wb[sheet].delete_rows(row_i, 1)

    wb.save(wb_new_name)

# #照片库ID list和字典
dic_={}
list_p_id=[]
for root ,dirs, files in pic_dirs:
        for f in files:
            dic_[f]=root
for key,value in dic_.items():
    list_p_id.append(key)

for root ,dirs, files in excel_dirs:
    for f in files:
        excel_path =os.path.join(root,f)
        new_f = f.replace(".xlsx", "_pic.xlsx")
        excel_new_path =os.path.join(aim_path,new_f)
        app = xw.App(visible=True, add_book=False)
        wb_pic = app.books.open(excel_path)
        ws=wb_pic.sheets["Sheet1"]
        rng =ws.range("A1").expand("table")
        max_row = rng.rows.count
        pic_id=1
        for row in range(2,max_row+1):
            id=str(ws.range(f"S{row}").value)+"__"+str(ws.range(f"T{row}").value)+"."
            for i in list_p_id:
                if id in i :
                    pic_address=os.path.join(dic_[i],i)
                    print(id, os.path.join(dic_[i],i),pic_id)
                    pic = ws.pictures.add(pic_address,\
                                            left=ws.range(f"U{row}").left, \
                                            top=ws.range(f"U{row}").top,\
                                            width=80, height=80)
                    pic.api.Placement=1
                    ws.range(f"U{row}").value = pic_id
                    pic_id += 1
        wb_pic.save(excel_new_path)
        wb_pic.close()
        app.quit()
if __name__ == '__main__':
    pass