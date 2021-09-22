
# encoding: utf-8
import PIL
import xwlings as xw
import sys,os,openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
excel_path=r"E:\work\samsung\test\workarea\W43 MS AuditChecklist-20191018.xlsx"
excel_path1=r"E:\work\samsung\test\workarea\W44 MS Audit Checklist-20191025更新.xlsx"
# excel_new =r"E:\work\samsung\W47周\W47 MS AuditChecklist-20191115 - 副本2gx.xlsx"
wb = openpyxl.load_workbook(excel_path)
wb1= openpyxl.load_workbook(excel_path1)
list_=[wb,wb1]
sheets = wb.sheetnames
def FuncSheet(InSheetNames,WbList):
    dic_item_value={}
    dic_item ={}
    list_item =[]
    for SheetName in InSheetNames:
        for i in range(len(WbList)):
            wb =WbList[i]
            ws =wb[SheetName]
            max_row =ws.max_row
            list_item0=[]
            list_item1=[]
            for row in range(3,max_row):
                in_value    = str(ws[f"B{row}"].value)
                test_value  = in_value[:3]
                test_value2 = ws[f"C{row}"].value
                test_value3 = ws[f"D{row}"].value
                if "." in test_value and test_value2:
                    list_item.append(in_value[3:])
                    list_item0.append({in_value[3:]:test_value2})
                    list_item1.append(in_value[3:])
            KeyId = SheetName +str(i)
            dic_item_value[KeyId]=list_item0
            dic_item [KeyId]=list_item1
    return(dic_item_value,dic_item,list_item)
ad = FuncSheet(sheets,list_)
def SoloListItem(InSheetNames,list_item,dic_item,WbList):
    list_sole_item=[]
    mid_list=[]
    for n in range(0,len(list_item)+1):
        for s in InSheetNames:
            for i in range(len(WbList)):
                KeyId = s +str(i)
                if n <len(dic_item[KeyId]):
                    mid_list.append(dic_item[KeyId][n])
        for i in mid_list:
            if i not in list_sole_item:
                list_sole_item.append(i)
    return list_sole_item
ds = SoloListItem(sheets,ad[2],ad[1],WbList=list_)
# print(ds,len(ds))

#插入列/行
# ws = wb.worksheets[0]
# ws.insert_cols(3)
# ws.insert_rows(14)

    
# ws1=wb.create_sheet("endsheet",0)
# r=3
# for k,v in dic_item_value.items():
#     #写入门店类型
#     ft1 = Font(color=colors.BLACK,name="微软雅黑")
#     ws1[f"B{r}"].value=k   
#     ws1[f"B{r}"].font =ft1
#     #写入 item 和 value 
#     for i in range(len(list_sole_item)):
#         ws1[f"{get_column_letter(i+3)}2"].value =list_sole_item[i]
        
#         ws1[f"{get_column_letter(i+3)}2"].font =ft1
#         for v in dic_item_value[ws1[f"B{r}"].value]:
#             for k,va in v.items():
#                 if k == list_sole_item[i]:
#                     ws1[f"{get_column_letter(i+3)}{r}"].value = va
#                     ft = Font(color=colors.RED,name="微软雅黑")
#                     if (r%2) == 0:
#                         ws1[f"{get_column_letter(i+3)}{r}"].font = ft
#                     else:
#                         ws1[f"{get_column_letter(i+3)}{r}"].font = ft1


#     ws1[f"A{r}"]=f"=sum(C{r}:{get_column_letter(len(list_sole_item)+2)}{r})"
#     r = r +1
    # ws1[f"{get_column_letter(i+3)}3"].value =dic_item_value[ws1["B3"].value][list_sole_item[i]]
    # ws1[f""]
# for key,values in dic_item_value.items():
#     print(values)
# print(dic_,len(dic_["新形象门店 (SES)"]))
excel_new = r"E:\work\samsung\test\workarea\end.xlsx"
excel_new1 = r"E:\work\samsung\test\workarea\end1.xlsx"
wb.save(excel_new)
wb1.save(excel_new1)
if __name__ == '__main__':
    pass
