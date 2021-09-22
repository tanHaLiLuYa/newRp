import os
import openpyxl
import xlwings as xw
from openpyxl.utils import get_column_letter, column_index_from_string

root_path = r"C:\Users\16930\Desktop"
aim_path = r"C:\Users\16930\Desktop\新建文件夹"
file_name = r"最新导购员照片表_李波_pic.xlsx"
split_by = "省"
file_path = os.path.join(root_path, file_name)
save_path =  os.path.join(aim_path, file_name)

# 查找分列依据在哪个列的函数


def get_cell_by_content(worksheet, content):
    for row in range(1, 5):
        for col in range(1, 50):
            if worksheet.range(f"{col}"f"{row}").value == content:
                if worksheet.range(row, col).value != "":
                    return [row, col]
                else:
                    return [row+1, col]


# 打开原始表
app = xw.App(visible=True, add_book=False)
wb =app.books.open(file_path)
# o_wb = openpyxl.load_workbook(file_path, data_only=True)

ws_active = o_wb.active
content_cell = get_cell_by_content(ws_active, split_by)
print(content_cell)
max_row = ws_active.max_row
col_name = get_column_letter(content_cell[1])
row_min = content_cell[0]+1

# 获取第一个sheet的分表依据的列表
split_list = []
for cell in ws_active[f"{col_name}{row_min}:{col_name}{max_row}"]:
    if cell[0].value != None:
        split_list.append(cell[0].value)

split_list = list(set(split_list))
print(split_list)

# 遍历要分表的依据，然后把原始文件另存为一个一个新文件
for l in split_list:
    wb_new_name = save_path.replace(".xlsx", f"_{l}.xlsx")  # 新文件名

    print(wb_new_name)
    o_wb.save(wb_new_name)
    wb = openpyxl.load_workbook(wb_new_name)
    sheets = wb.sheetnames
    # 遍历新文件的各个sheet：
    for sheet in sheets:

        cell_position = get_cell_by_content(
            wb[sheet], split_by)  # 这是字符串所在的那个格的位置
        col_name = get_column_letter(cell_position[1])  # 这是那个格的列号

        print(f"在《{sheet}》里，「{split_by}」所在的位置为{cell_position}")
        row_max = wb[sheet].max_row  # 这是最大行号
        # 这是总部编码下面的那个行
        row_min = cell_position[0]

        print(f"{split_by}下面有字的第一行是{row_min}")
        

        for row_i in range(row_max, row_min, -1):
            v = wb[sheet][f"{col_name}{row_i}"].value
            if v == l:
                pass
            else:
                # print("将要删除这一行")
                
                wb[sheet].delete_rows(row_i, 1)

                if wb[sheet][f"U{row_i}"].value:
                    del wb[sheet]._images[int(wb[sheet] [f"U{row_i}"].value)]

    wb.save(wb_new_name)


