#!/usr/bin/env python
# encoding: utf-8
import os
import sys
import datetime
import time

from openpyxl import load_workbook,Workbook

root_path = r"D:\work\samsung\W19周报\异常"
our_file = os.path.join(
    root_path, "【W19】导购员异常和门店异常情况_CBIC_0513_1557.xlsx")
our_back_file = our_file.replace(".xlsx","_backup.xlsx")
sx_D = os.path.join(root_path, "离岗报备记录导出-D1118.xlsx")
sx_A = os.path.join(root_path, "离岗信息-A1118.xlsx")
sx_F = os.path.join(root_path, "全国特殊事件数据-F1118.xlsx")
sx_P = os.path.join(root_path,"导购员排班表W19-5.12.xlsx")
id_our_col = "I"
name_our_col = "J"
date_our_col = "K"
time_in_col = "L"
time_out_col = "M"
#获取周几
def get_week_day(date):
    week_day_dict = {0:"星期一",1:"星期二",2:"星期三",3:"星期四",4:"星期五",5:"星期六",6:"星期天"}
    day = date.weekday()
    return week_day_dict[day]

# print(our_file,"\n", sx_A, "\n", sx_D, "\n", sx_F)

#判断检查时间与报备时间是否重合，1为重合
def time_overlap(start1, end1, start2, end2):
    if (start1 <= end2) and (start2 <= end1):
        return 1
    else:
        return 0
#判断进店时间是否不在上班时间内，1为不在上班时间里面
def time_overlap1(start1, end1, start2, end2):
    if (start1 >= end2) or (start2 >= end1):
        return 1
    else:
        return 0
#备份函数
def backup(in_file=our_file, outfile=our_back_file):
    wb_our = load_workbook(our_file)
    # wb_our.remove(["异常导购员List"])
    wb_our.save(our_back_file)

#事后修改排班表
def shihouxiugai(our_file=our_file,sxfile=sx_P):
    wb_our = load_workbook(our_file)
    ws_our = wb_our["异常导购员List"]
    wb_sx = load_workbook(sxfile)
    ws_sx = wb_sx.active
    ws_our["AK1"].value = "事后修改排班_判断"
    ws_our["AL1"].value = "上班时间"
    ws_our["AM1"].value = "下班时间"

    weekday_sx_dict = {"星期三":"L","星期四":"N","星期五":"P","星期六":"R","星期天":"T","星期一":"V","星期二":"X"}
    #往右加一列
    weekday_sx_dict2 = {"星期三":"M","星期四":"O","星期五":"Q","星期六":"S","星期天":"U","星期一":"W","星期二":"Y"}
    #导购员ID_list
    id_sx_list =[]

    # E列为导购员ID，L列开始显示日期列
    for row in range(2, ws_our.max_row + 1):
    
        id_our = ws_our[f"{id_our_col}{row}"].value
        date_our = ws_our[f"{date_our_col}{row}"].value
        time_in = ws_our[f"{time_in_col}{row}"].value
        time_out = ws_our[f"{time_out_col}{row}"].value
        time_start=str(date_our)[:11]  + str(time_in)[:5]
        our_start = datetime.datetime.strptime(f"{time_start}", "%Y-%m-%d %H:%M")

        time_end=str(date_our)[:11] + str(time_out)[:5]
        our_end = datetime.datetime.strptime(f"{time_end}", "%Y-%m-%d %H:%M")
        # time_begin =""
        # time_end = ""
        # recording = ""

        for row2 in range(3,ws_sx.max_row+1):
            id_sx = ws_sx[f"E{row2}"].value#导购员ID列
            id_sx_list.append(id_sx)
            if id_sx == id_our:
                column_time = weekday_sx_dict[get_week_day(our_start)]
                column_time2 = weekday_sx_dict2[get_week_day(our_start)]
                time_begin_str=ws_sx[f"{column_time}{row2}"].value
                time_end_str = ws_sx[f"{column_time2}{row2}"].value
                time_work_begin=str(date_our)[:11] +str(time_begin_str)[:5]
                time_work_end = str(date_our)[:11] +str(time_end_str)[:5]
                # print(our_start, our_end, time_work_begin,time_work_end)
                # # time_begin = datetime.datetime.combine(date_our, time_begin_str)
                # # time_end = datetime.datetime.combine(date_our, time_end_str)
                # time_begin = datetime.datetime.strptime(f"{time_work_begin}", "%Y-%m-%d %H:%M")
                # time_end = datetime.datetime.strptime(f"{time_work_end}", "%Y-%m-%d %H:%M")
                # # print(column_time)
                if ws_sx[f"{column_time}{row2}"].value  in ["休息" ,"病假","周休","调休","离职","事假","产假","闭店","婚假","休班","其他"]:
                    ws_our[f"AK{row}"].value = "提前修改排班表"
                    ws_our[f"AL{row}"].value = ws_sx[f"{column_time}{row2}"].value
                    ws_our[f"AM{row}"].value = ws_sx[f"{column_time}{row2}"].value
                elif time_overlap1(our_start, our_end, datetime.datetime.strptime(f"{time_work_begin}", "%Y-%m-%d %H:%M")
                    , datetime.datetime.strptime(f"{time_work_end}", "%Y-%m-%d %H:%M")) == 1:
                    ws_our[f"AK{row}"].value = "提前修改排班表"
                    ws_our[f"AL{row}"].value = str(datetime.datetime.strptime(f"{time_work_begin}", "%Y-%m-%d %H:%M"))+get_week_day(our_start)
                    ws_our[f"AM{row}"].value = str(datetime.datetime.strptime(f"{time_work_end}", "%Y-%m-%d %H:%M"))+get_week_day(our_start)
        if id_our not in id_sx_list:
            ws_our[f"AK{row}"].value = "没有排班表"
            # print("匹配没有排班表导购员成功",row2)
            # else:
            #     ws_our[f"AK{row}"].value = "没有排班表"
                
    wb_our.save(our_file)
                
# 离岗报备——D 记录导出表的函数：
def ligangbaobei(our_file=our_file, sxfile=sx_D):
    wb_our = load_workbook(our_file)
    ws_our = wb_our["异常导购员List"]
    wb_sx = load_workbook(sxfile)
    ws_sx = wb_sx.active
    ws_our["AE1"].value = "离岗报备_判断"
    ws_our["AF1"].value = "离岗报备_详情"

    for row in range(2, ws_our.max_row + 1):

        id_our = ws_our[f"{id_our_col}{row}"].value
        date_our = ws_our[f"{date_our_col}{row}"].value
        time_in = ws_our[f"{time_in_col}{row}"].value
        time_out = ws_our[f"{time_out_col}{row}"].value
        
        time_start=str(date_our)[:11] + str(time_in)[:5]
        our_start = datetime.datetime.strptime(f"{time_start}", "%Y-%m-%d %H:%M")

        time_end=str(date_our)[:11] + str(time_out)[:5]
        our_end = datetime.datetime.strptime(f"{time_end}", "%Y-%m-%d %H:%M")
        # print(">>>>>",our_start, our_end)
        baobei = ""
        for r2 in range(3, ws_sx.max_row + 1):
            id_sx = ws_sx[f"A{r2}"].value#+++++++++++++++需要确认是否发生列变化  A

            if id_sx == id_our:
                
                date_2 = ws_sx[f"C{r2}"].value#+++++++++++++++需要确认是否发生列变化   C D E 
                time_begin_str = ws_sx[f"D{r2}"].value if len(ws_sx[f"D{r2}"].value) >=5 else "08:00"
                time_end_str = ws_sx[f"E{r2}"].value if len(
                    ws_sx[f"E{r2}"].value) >= 5 else "23:59"
                time_begin = datetime.datetime.strptime(f"{date_2} {time_begin_str}", "%Y-%m-%d %H:%M")
                time_end = datetime.datetime.strptime(f"{date_2} {time_end_str}", "%Y-%m-%d %H:%M")
                print("匹配离岗报备文件成功",row,r2,)
                if time_overlap(our_start, our_end, time_begin, time_end) == 1:
                    
                    note = f"《离岗报备记录表》 中第{r2}行，报备离岗时间为{time_begin},返岗时间为{time_end}\n"
                    baobei = baobei + note
                    ws_our[f"AE{row}"].value = "有时间重合"
                    ws_our[f"AF{row}"].value = baobei
                    print("+++++", baobei)
    wb_our.save(our_file)
                
#对离岗信息文件匹配
def ligangxinxi(our_file=our_file, sxfile=sx_A):
    wb_our = load_workbook(our_file)
    ws_our = wb_our["异常导购员List"]
    wb_sx = load_workbook(sxfile)
    ws_sx = wb_sx.active
    ws_our["AG1"]="离岗信息_判断"
    ws_our["AH1"]="离岗信息_详情"
    for row in range(2, ws_our.max_row + 1):

        id_our = ws_our[f"{id_our_col}{row}"].value
        date_our = ws_our[f"{date_our_col}{row}"].value
        time_in = ws_our[f"{time_in_col}{row}"].value
        time_out = ws_our[f"{time_out_col}{row}"].value
        
        time_start=str(date_our)[:11] + str(time_in)[:5]
        our_start = datetime.datetime.strptime(f"{time_start}", "%Y-%m-%d %H:%M")

        time_end=str(date_our)[:11] + str(time_out)[:5]
        our_end = datetime.datetime.strptime(f"{time_end}", "%Y-%m-%d %H:%M")
        # print(">>>>>",our_start, our_end)
        baobei = ""
        for r2 in range(2, ws_sx.max_row + 1):
            id_sx = ws_sx[f"N{r2}"].value#+++++++++++++++需要确认是否发生列变化 N         

            if id_sx == id_our:

                date_2 = ws_sx[f"O{r2}"].value#+++++++++++++++需要确认是否发生列变化   O P Q
                time_begin_str = ws_sx[f"P{r2}"].value if len(
                    ws_sx[f"P{r2}"].value) >= 5 else f"{date_2} 09:00"
                time_end_str = ws_sx[f"Q{r2}"].value if len(
                    ws_sx[f"Q{r2}"].value) >= 5 else f"{date_2} 23:59"
                time_begin = datetime.datetime.strptime( time_begin_str, "%Y/%m/%d %H:%M")
                time_end = datetime.datetime.strptime( time_end_str, "%Y/%m/%d %H:%M")
                print("匹对离岗信息文件成功",row, r2,)
                if time_overlap(our_start, our_end, time_begin, time_end) == 1:

                    note = f"《离岗信息》中第{r2}行，报备离岗时间为{time_begin},返岗时间为{time_end}\n"
                    baobei = baobei + note
                    ws_our[f"AG{row}"].value = "有时间重合"
                    ws_our[f"AH{row}"].value = baobei
                    print("+++++", baobei)
    wb_our.save(our_file)

#特殊事件文件匹对
def teshushijian(our_file=our_file, sxfile=sx_F):
    wb_our = load_workbook(our_file)
    ws_our = wb_our["异常导购员List"]
    wb_sx = load_workbook(sxfile)
    ws_sx = wb_sx.active
    ws_our["AI1"]="特殊事件_判断"
    ws_our["AJ1"]="特殊事件_详情"

    for row in range(2, ws_our.max_row + 1):
        name_our = ws_our[f"{name_our_col}{row}"].value
        date_our = ws_our[f"{date_our_col}{row}"].value
        time_in = ws_our[f"{time_in_col}{row}"].value
        time_out = ws_our[f"{time_out_col}{row}"].value
        
        time_start=str(date_our)[:11] + str(time_in)[:5]
        our_start = datetime.datetime.strptime(f"{time_start}", "%Y-%m-%d %H:%M")

        time_end=str(date_our)[:11] + str(time_out)[:5]
        our_end = datetime.datetime.strptime(f"{time_end}", "%Y-%m-%d %H:%M")
        # print(">>>>>",our_start, our_end)
        baobei = ""
        for r2 in range(2, ws_sx.max_row + 1):
            name_sx = ws_sx[f"A{r2}"].value#+++++++++++++++需要确认是否发生列变化  A C
            section = ws_sx[f"C{r2}"].value
            if name_sx == name_our:
                
                time_begin_str = ws_sx[f"D{r2}"].value#+++++++++++++++需要确认是否发生列变化   D E
                time_end_str = ws_sx[f"E{r2}"].value
                time_begin = datetime.datetime.strptime(time_begin_str, "%Y-%m-%d %H:%M")
                time_end = datetime.datetime.strptime(time_end_str, "%Y-%m-%d %H:%M")
                print("匹对特殊事件成功",row, r2,)
                if time_overlap(our_start, our_end, time_begin, time_end) == 1:

                    note = f"《特殊事件数据》中第{r2}行，姓名{name_sx}，部门为{section}，报备离岗时间为{time_begin},返岗时间为{time_end}\n"
                    baobei = baobei + note
                    ws_our[f"AI{row}"].value = "有时间重合"
                    ws_our[f"AJ{row}"].value = baobei
                    print("+++++", baobei)
    wb_our.save(our_file)

# backup()
shihouxiugai()
# ligangbaobei()
# ligangxinxi()
# teshushijian()
