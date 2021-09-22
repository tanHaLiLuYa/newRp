import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
os.chdir(r"D:\work\分表")  # 固定输入 和输出文件位置


class Solution:

    def __init__(self):
        self.gist = input("分列依据：")
        self.fileName = input("文件名称（带.后缀）：")

    def get_cell_by_content(self, workSheet, content):
        # 从前5行 和前50列 范围内查找
        for row in range(1, 5):
            for col in range(1, 50):
                if workSheet.cell(row, col).value == content:
                    if workSheet.cell(row+1, col).value != None:
                        return [row, col]
                    else:
                        return [row+1, col]

    def run(self):
        # 打开原始表
        o_wb = openpyxl.load_workbook(self.fileName, data_only=True)
        ws_active = o_wb.active
        content_cell = self.get_cell_by_content(ws_active, self.gist)
        print("{}的单元格坐标为：".format(self.gist),content_cell)
        max_row = ws_active.max_row
        col_name = get_column_letter(content_cell[1])
        row_min = content_cell[0]+1
        # 获取第一个sheet的分表依据的列表
        split_list = []
        for cell in ws_active[f"{col_name}{row_min}:{col_name}{max_row}"]:
            if cell[0].value != None:
                split_list.append(cell[0].value)

        split_list = list(set(split_list))#去重
        print(split_list)
        # 遍历要分表的依据，然后把原始文件另存为一个一个新文件
        for l in split_list:
            wb_new_name = self.fileName.replace(".xlsx", f"_{l}.xlsx")  # 新文件名

            print(wb_new_name)
            o_wb.save(wb_new_name)
            wb = openpyxl.load_workbook(wb_new_name)
            sheets = wb.sheetnames
            # 遍历新文件的各个sheet：
            for sheet in sheets:

                cell_position = self.get_cell_by_content(
                    wb[sheet], self.gist)  # 这是字符串所在的那个格的位置
                col_name = get_column_letter(cell_position[1])  # 这是那个格的列号

                print(f"在《{sheet}》里，「{self.gist}」所在的位置为{cell_position}")
                row_max = wb[sheet].max_row  # 这是最大行号
                # 这是总部编码下面的那个行
                row_min = cell_position[0]

                print(f"{self.gist}下面有字的第一行是{row_min}")

                for row_i in range(row_max, row_min, -1):  # 从后往前遍历
                    v = wb[sheet][f"{col_name}{row_i}"].value
                    if v == l:
                        pass
                    else:
                        wb[sheet].delete_rows(row_i, 1)
                print("*"*20)
            wb.save(wb_new_name)


if __name__ == "__main__":
    tet = Solution()
    tet.run()
