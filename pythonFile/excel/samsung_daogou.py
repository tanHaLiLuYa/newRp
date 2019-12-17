#!/usr/bin/env python
# encoding: utf-8
import os
import sys
import datetime
import time

# 首先是周数的列表
week_list = ["W19"]
# week_list = ["W02", "W03", "W08", "W09", "W10", "W11", "W13", "W14", "W15", "W16", "W17", "W19"]
for week in week_list:
    print(week)

from openpyxl import load_workbook
dg_sheet = "/Users/simon/Desktop/三星导购员/W02_W19.xlsx"


def get_dg_status(week):

    wb_dg = load_workbook(filename=dg_sheet)
    ws_dg = wb_dg['data']
    pb_sheet = f"/Users/simon/Desktop/三星导购员/{week}导购员排班表.xlsx"
    ws_pb = load_workbook(filename=pb_sheet).active
    print(ws_pb.title)

    def get_date_col(visit_date):
        for row in ws_pb["A1":"AZ1"]:
            for d in row:
                if d.value == visit_date:
                    return [d.column, d.offset(row=0, column=1).column]
                    

    for i in range(2, ws_dg.max_row + 1):
        week_value = ws_dg[f"B{i}"].value
        if week_value != week:
            pass
        else:
            ap_id = ws_dg[f"C{i}"].value
            ap_date = ws_dg[f"E{i}"].value
            # ap_datetime = datetime.datetime.combine(ap_date, ws_dg[f"F{i}"].value)
            # print(str(i),ap_date.year, ap_date.month, ap_date.day)

            # j 代表排班表里的行数
            for j in range(3, ws_pb.max_row + 1):
                print("导购行", i, "---排班表行", j, "--------")
                p_id = ws_pb[f'D{j}'].value
                if week_value == week and ap_id == p_id:
                    print("导购行", i, "---排班表行", j, "--------", get_date_col(ap_date))
                    
                    time_col_start = get_date_col(ap_date)[0]
                    time_col_end = get_date_col(ap_date)[1]
                    time_start_cell = ws_pb[time_col_start + f"{j}"].value
                    time_end_cell = ws_pb[time_col_end + f"{j}"].value
                    if time_start_cell == "休假":
                        ws_dg[f"H{i}"].value = "休假"
                        ws_dg[f"I{i}"].value = "休假"
                        ws_dg[f"J{i}"].value = "休假中"
                    else:             
                        time_start_str = f"{ap_date.year}-{ap_date.month}-{ap_date.day} {time_start_cell}"
                        time_end_str = f"{ap_date.year}-{ap_date.month}-{ap_date.day} {time_end_cell}"
                        print(time_start_str, time_end_str)
                        time_start_value = datetime.datetime.strptime(
                            time_start_str, "%Y-%m-%d %H:%M")
                        time_end_value = datetime.datetime.strptime(
                            time_end_str, "%Y-%m-%d %H:%M")
                        ap_datetime = datetime.datetime.combine(ap_date, ws_dg[f"F{i}"].value)
                        if ap_datetime > time_start_value and ap_datetime < time_end_value:
                            print(f"{i}",ap_datetime, time_start_value, time_end_value, "----", True)
                            ws_dg[f"H{i}"].value = time_start_value
                            ws_dg[f"I{i}"].value = time_end_value
                            ws_dg[f"J{i}"].value = "导购员在上班"
                        else:
                            print(f"{i}", ap_datetime, time_start_value,time_end_value, "<<<<<<", False)
                            ws_dg[f"H{i}"].value = time_start_value
                            ws_dg[f"I{i}"].value = time_end_value
                            ws_dg[f"J{i}"].value = "访问时间不在上班时间内"
    wb_dg.save(dg_sheet)
    print(f"{week}期的导购员上班时间已经写完")


class Main():
    def __init__(self):
        pass


if __name__ == '__main__':
    for w in week_list:
        get_dg_status(w)
