import os
import re
import sys
import datetime
import time
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook, Workbook


def strToDicForTab(strIn):
    strIn = str(strIn)
    a = {}
    a["T720"] = int (strIn[strIn.find("T720")+4]) if strIn.find("T720") != -1 else 0
    a["T510"] = int (strIn[strIn.find("T510")+4]) if strIn.find("T510") != -1 else 0
    a["T860"] = int (strIn[strIn.find("T860")+4]) if strIn.find("T860") != -1 else 0
    a["P610"] = int (strIn[strIn.find("P610")+4]) if strIn.find("P610") != -1 else 0
    return a
def tabIsSame(dic, dic_):
    for i in ["P610", "T860", "T510", "T720"]:
        if int(dic[i]) != int(dic_[i]):
            return False
    return True


root_path = r"D:\work\samsung\汇总\tab"  # 目录

os.chdir(root_path)#更改目录

fileName = r"Tab覆盖门店检查明细 -核对1130_1730.xlsx"  # 文件名称

wb = load_workbook(fileName)

ws2 = wb.active


for r in range(2,ws2.max_row+1):
    # print(ws2["T{}".format(r)].value)
    print(r)
    dict1 = strToDicForTab(ws2["K{}".format(r)].value)
    dict2 = strToDicForTab(ws2["L{}".format(r)].value)
    # print(dict2,r)
    if not tabIsSame(dict1,dict2) :
        if sum(dict1.values()) == sum(dict2.values()):
            ws2["M{}".format(r)].value="型号不一致"
        elif sum(dict1.values()) < sum(dict2.values()):
            ws2["M{}".format(r)].value="小于系统"
        else:
            ws2["M{}".format(r)].value="大于系统"
else:
    print("完成核对。。。")
    

fileNameNew = "Tab覆盖门店检查明细 -核对1130_1730.xlsx"
wb.save(fileNameNew)
wb.close()

