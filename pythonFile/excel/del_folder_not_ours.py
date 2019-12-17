#!/usr/bin/env python
# encoding: utf-8
from openpyxl import load_workbook
import os
import re
import shutil

'''
实现功能：
有一个excel表是我们的店面，有一个文件夹里面有很多文件夹，只想保留文件夹中属于我们的店面

思路：
1、把excel表中的店面ID读成一个列表；
2、遍历文件夹
3、如果这个文件夹的名字不在列表中，删除

'''

excel_path = "/Users/simon/Documents/12月消费自检照片导出/ours.xlsx"
folder_path = "/Users/simon/Documents/12月消费自检照片导出/"

# 读取为列表
wb = load_workbook(filename=excel_path)
ws = wb.get_sheet_by_name('ours')
print(wb.get_sheet_names())
cells = ws["D2":"D1785"]
our_list = []
for row in cells:
    for cell in row:
        our_list.append(cell.value)

print(our_list)

# 读取文件夹

list_dirs = os.walk(folder_path)

for root, folders, files in list_dirs:
    for folder in folders:
        if folder[:8] not in our_list:
            #为啥加个8哩，因为啊有些后面有那个（1）（2）啊啥的。只对比文件夹的前8位就可以了。
            print(folder[:8], "不在我们的列表中,开始删除")
            del_path = os.path.join(root, folder)
            shutil.rmtree(del_path)
            print("已经删除", del_path)
        else:
            print(folder,"在列表中，需要保留")
