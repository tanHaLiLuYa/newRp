import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter



def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a
def namematch(str1,str2):
    name1= "".join(re.findall(namepat,str1))
    list2= re.findall(namepat,str2)
    if name1 not in list2:
        return False , name1
    else :
        return True , name1



def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)



def otherValue(letter,r, num):
    tNum = column_index_from_string(letter)
    letternew= get_column_letter(tNum+num)
    return ws2["{}{}".format(letternew,r)].value



print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"E:\work\tpp\samsung\2021年\06月\W26\洗库"


os.chdir(root_path)

fileName = "【W26】【洗库】0627_1750.xlsx"  

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = re.compile(r"[A-Z]{1,2}[0-9]{1,3}.{0,1}[0-9]{0,1}[.]")
matchPatternID =re.compile(r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}") 
namepat = re.compile(r"[\u4e00-\u9fa5]{1,12}")

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.copy_worksheet(wb["提交"])
ws2.title = "check1"


queDic = {}
titleDic = {}

# 开始列 X,结束列 ,问题记录成字典内的列表
starColumnName = "X"
endColumnName = "DK"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
    if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
        titleDic["{}".format(get_column_letter(c))] = re.search(
            matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    queDic["C{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)
        print("列：", testCol, "行：", rowNumber)
        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"

        investorType = ws2["K{}".format(rowNumber)].value
        investorType = investorType[0:3]  # 投资类型
        # imageType = ws2["T{}".format(rowNumber)].value
        # imageType = imageType[0:1]  # 新旧形象

        datiren = ws2["CK{}".format(rowNumber)].value if ws2["CK{}".format(rowNumber)].value else "none"
        datiren =datiren[0:1]
        # ————————————————————————————————————————————————————————————————————————————

        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        # 身份证
        if testCol == "X":
            testIDList = otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                queDic["C{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                                                       format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))

        # 1 门店形象 店外露出
        # ————————————————————————————————————————————————————————————————————————————

        if testCol in ["Z", "AA","AB","BM"] and testValue[0:1]  not in ["1", "N"]  :  #, 
            fillYorN = True

        if testCol in ["AE"] and testValue[0:1] not in ["1", "4", "N"] :
            fillYorN = True
        if testCol in ["AC","AD"] and testValue[0:1] not in ["1", "3", "N"] :
            fillYorN = True

        # 体验台销售柜
        if testCol in ["AH", "AK", "AN", "AQ"] and testValue != "NA":
            touru = otherValue(testCol,rowNumber, -1) if otherValue(testCol,rowNumber, -1) else 0
            fillYorN = True if int(touru) != int(testValue) else False
        # F 52 机模、真机陈列
        if ws2["BA{}".format(rowNumber)].value=="Y":
            if testCol in ["AZ","AY"] and  testValue[0:1] not in ["1", "N"] :
                fillYorN = True
        # # 3D展架 陈列
        # if ws2["BC{}".format(rowNumber)].value=="Y":
        #     if testCol in ["BA","BB"] and  testValue[0:1] != "1" :
        #         fillYorN = True
        #S21 真机、露出
        if testCol =="BC" and otherValue(testCol,rowNumber, -10)=="Y" and testValue[0:1] not in [ "N","2"]:
            fillYorN = True
        #A52
        if ws2["AU{}".format(rowNumber)].value=="Y":
            if testCol in ["BF","BE"] and  testValue[0:1] not in [ "N","1"] :
                fillYorN = True

        # #促销
        # if testCol in ["BE","BF"] and testValue[0:1] not in ["1", "3"] and investorType=="SES" and imageType !="3" :
        #     fillYorN = True

        #电子屏
        if ws2["BI{}".format(rowNumber)].value >0:
            if testCol in ["BH"] and  testValue[0:1] not in ["1", "N"]:
                fillYorN = True
        #xun 2 d     
        if testCol in [ "BK","BJ","BW","BX","BY"] and testValue[0:1]  =="2"  :  #, 
            fillYorN = True    
        if testCol in ["BZ","CA","CB"]and  testValue[0:1] not in ["1", "N"]:
            fillYorN = True

        #F52 销售柜
        if ws2["BA{}".format(rowNumber)].value =="Y": 
            if testCol in [ "BP","BT"]  and testValue[0:1] not in [ "N","1","4"]:
                fillYorN = True      
            if testCol =="BQ"  and testValue[0:1] not in [ "N","1","5"]:
                fillYorN = True   
            if testCol in [ "BS","BR"] and testValue[0:1]  =="2"  :  #, 
                fillYorN = True 
        #A52
        if ws2["AU{}".format(rowNumber)].value =="Y":
            if testCol in ["BU","BV"] and  testValue[0:1] not in ["1", "N"]:
                fillYorN = True
  
       #顺电
        if ws2["CC{}".format(rowNumber)].value =="Y":
            if testCol in ["CD","CE"] and  testValue[0:1] not in ["1", "N"]:
                fillYorN = True
            if ws2["BI{}".format(rowNumber)].value >0:
                if testCol in ["CF","CG","CH"] and  testValue[0:1] not in ["1", "N"]:
                    fillYorN = True
        #ren yuan 
        if testCol=="CJ" and testValue[0:1] in ["3","5"]:
            fillYorN = True
        
         
        #答题人姓名 正确与否
        if testCol == "CL" and otherValue(testCol,rowNumber,-1)[0:1] =="1":
            testp = otherValue(testCol,rowNumber,-65) if otherValue(testCol,rowNumber,-65)  else "无"
            YesOrNo ,name = namematch(testValue,testp)
            ws2["CL{}".format(rowNumber)].value = name
            if not YesOrNo:
                fillYorN = True 
        #知识考核  多选题 == -1 是无该选项 ， 多项选择判断条件成立则为 错误答案
        if datiren=="1":
            if testCol =="CR":
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True
       
            if testCol =="CM":
                fillYorN = True if testValue.find("隔空操作") == -1 else False 
            if testCol =="CN":
                fillYorN = True if testValue.find("D、") == -1 else False 
            if testCol =="CO":
                fillYorN = True if testValue.find("C、") == -1 else False 
            if testCol =="CP":
                fillYorN = True if testValue.find("A、") == -1 else False 
            if testCol =="CQ":
                fillYorN = True if testValue.find("B、") == -1 else False 
            if testCol =="CS":
                fillYorN = True if testValue.find("B、") == -1 else False 
            if testCol =="CT" and investorType=="SIS":
                fillYorN = True if testValue.find("A、") == -1 else False 

        #竞品 CU 2
        if testCol in ["CU","DE","DF","DG","DH","DI"] and testValue[0:1] =="2":
            fillYorN = True
        
        # 违规  
        # if testCol == "CC" and testValue[0:2]not in ["1.", "2.", "3."]:
        #     fillYorN = True
        # if testCol == "DK" and testValue[0:2]not in ["7.", "8.", "9."]:
        #     fillYorN = True
        
        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            queDic["C{}".format(rowNumber)].append(
                "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
for r in range(2, ws2.max_row+1):
    valueIn = " ".join(queDic["C{}".format(r)])
    ws2["C{}".format(r)].value = valueIn

# 保存修改

print("saving......")

wb.save(fileName.replace(".xlsx","-修改.xlsx"))

wb.close()
