import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter



def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a
def namematch(str1,str2):
    name1= "".join(re.findall(namepat,str1))
    list2= re.findall(namepat,str2)
    if name1 not in list2:
        return False , name1
    else :
        return True , name1
def N20YANGJI(list1, list2):
    if len(set(list1)) != len(list1) or len(list2) % 2 != 0:
        return True
    else:
        listTest = []
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])

        for i in listTest:
            if i[0].isdigit():
                if int(i[0]) > 2 or (int(i[0]) > 1 and i[1] != "金"):
                    return True
            else:
                if int(i[1]) > 2 or (int(i[1]) > 1 and i[0] != "金"):
                    return True
        return False


def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)

        # listTest = []
        # for i in range(0,len(list2),2):
        #     listTest.append(list2[i]+list2[i+1])
        # for i in listTest:
        #     if i[0].isdigit():


def strToDicForTab(strIn):
    a = {}
    a["T720"] = strIn[strIn.find("T720")+4] if strIn.find("T720") != -1 else 0
    a["T510"] = strIn[strIn.find("T510")+4] if strIn.find("T510") != -1 else 0
    a["T860"] = strIn[strIn.find("T860")+4] if strIn.find("T860") != -1 else 0
    a["P610"] = strIn[strIn.find("P610")+4] if strIn.find("P610") != -1 else 0

    return a


def testTabFun(dic, dic_):
    for i in ["P610", "T860", "T510", "T720"]:
        if int(dic[i]) != int(dic_[i]):
            return True
    return False


def otherValue(letter,r, num):
    tNum = column_index_from_string(letter)
    letternew= get_column_letter(tNum+num)
    return ws2["{}{}".format(letternew,r)].value



print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"E:\work\tpp\samsung\2021年\02月\W08\洗库"


os.chdir(root_path)

fileName = "【W09】【洗库】0223_1600.xlsx"  

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = re.compile(r"[A-W]{1,2}[0-9]{1,3}.{0,1}[0-9]{0,1}[.]")
matchPatternID =re.compile(r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}") 
namepat = re.compile(r"[\u4e00-\u9fa5]{1,12}")

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.copy_worksheet(wb["提交"])
ws2.title = "check1"


queDic = {}
titleDic = {}

# 开始列 X,结束列 ,问题记录成字典内的列表
starColumnName = "X"
endColumnName = "CW"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
    if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
        titleDic["{}".format(get_column_letter(c))] = re.search(
            matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    queDic["C{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)
        print("列：", testCol, "行：", rowNumber)
        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"

        investorType = ws2["K{}".format(rowNumber)].value
        investorType = investorType[0:3]  # 投资类型
        imageType = ws2["T{}".format(rowNumber)].value
        imageType = imageType[0:1]  # 新旧形象

        datiren = ws2["CF{}".format(rowNumber)].value if ws2["CF{}".format(rowNumber)].value else "none"
        datiren =datiren[0:1]
        # ————————————————————————————————————————————————————————————————————————————

        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        # 身份证
        if testCol == "X":
            testIDList = otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                queDic["C{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                                                       format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))

        # 1 门店形象 店外露出
        # ————————————————————————————————————————————————————————————————————————————

        if testCol in ["AB", "Z", "AA","AC","AX","AY","BB","BC","BX","CR"] and testValue[0:1] != "1" :  #, 
            fillYorN = True

        if testCol in ["AD"] and testValue[0:1] not in ["1", "4"] :
            fillYorN = True

        # 体验台销售柜
        if testCol in ["AG", "AJ", "AM", "AP"] :
            touru = otherValue(testCol,rowNumber, -1) if otherValue(testCol,rowNumber, -1) else 0
            fillYorN = True if int(touru) != int(testValue) else False
        
        #4.1.2 LFD（电视屏幕）是否有播放S21系列5G KV画面内容
        if testCol in ["AV"] and testValue[0:1] not in ["1", "3"] :
            fillYorN = True
        #4.3.1 S21系列5G机模或真机露出
        if testCol =="AW" and otherValue(testCol,rowNumber,-4)=="Y"  and testValue[0:1] != "1" :
            fillYorN = True 
        #4.1.3 S21系列 5G新品KV A4立牌露出 AX 1
        #4.1.4 S21系列价格签上"新品"  AY 1
        # 4.1.5 S21系列5G新品物料氛围明显
        if testCol  =="AZ" :
            if  testValue.find("9.") != -1:
                fillYorN = True 
            else:
                fillYorN = True  if testValue.count(".") < 4 else False

        #

        # 4.3.2 Buds pro机模或真机露出BB 1
        #4.3.3 Fit2 机模或真机露出 BC 1

        # 4.4.1 新品台陈列
        if testCol == "BD" and imageType in ["1","2"]  and investorType in ["SES","SIS"]  :
            fillYorN = True if testValue[0:1] not in ["1", "5"]  else False

        #4.4.1 新品区陈列
        if testCol =="BE" and testValue[0:1] not in ["1","3"] and imageType == "2" :  
            fillYorN = True
        # 4.7.1 新品台陈列规范
        if testCol == "BF" and imageType in ["1","2"]  and investorType in ["SES","SIS"]  :
            fillYorN = True if testValue[0:1] not in ["1", "4"]  else False

        #4.4.2 S21系列 5G精品陈列
        if testCol =="BG" and testValue[0:1] not in ["1"] and otherValue(testCol,rowNumber,1)=="是":  
            fillYorN = True
        #4.4.3 S21 Ultra 5G黑色陈列套装组
        if testCol =="BI" and imageType in ["1","2"] and  testValue[0:1] !="1" and otherValue(testCol,rowNumber,2)=="Y":  
            fillYorN = True        
        if testCol =="BJ" and imageType in ["3"] and  testValue[0:1] !="1" and otherValue(testCol,rowNumber,1)=="Y": 
            fillYorN = True  
        #4.4.4 S21+5G 紫色陈列套装组  4.4.5 S21 5G 陈列套装组
        if testCol in ["BL","BN"] and otherValue(testCol,rowNumber,1)=="Y" and  testValue[0:1] !="1":
            fillYorN = True  
        #4.6.1 S21系列5G手机保护壳露出
        if testCol =="BP" and testValue[0:1] not in ["1", "3"] :
            fillYorN = True   
        #E20. 店内有几色 S20 FE 露出？  
        if testCol =="BQ" and testValue[0:1] not in ["1"] and otherValue(testCol,rowNumber,-22)=="Y":  
            fillYorN = True   
        #4.6.2 配件桌陈列
        if testCol =="BR" and testValue[0:1] not in ["1", "3"]:
            fillYorN = True   
        #4.6.3 配件墙陈列
        if testCol =="BT" :
            numsV = otherValue(testCol,rowNumber,-1)[0:1]
            if numsV in ["1","3"]:
                fillYorN = True   if testValue.find("2.") == "-1" or testValue.count(".")>1 else False
            elif numsV=="2":
                fillYorN = True   if testValue.find("2.") == "-1" or testValue.find("3.") == "-1" or testValue.count(".")!=2 else False
            elif numsV=="4":
                fillYorN = True   if testValue.find("2.") == "-1" or testValue.find("3.") == "-1" or testValue.find("1.") == "-1"or testValue.count(".")!=3 else False
            elif numsV=="5":
                fillYorN = True   if testValue.find("2.") == "-1"  or testValue.find("1.") == "-1"or testValue.count(".")!=2 else False
            else:
                fillYorN = True   if testValue.find("5.") == "-1" and testValue.find("6.") == "-1" or testValue.count(".")!=1 else False
        if testCol =="BU" and testValue[0:1] not in ["1","5","6"] and  imageType in ["1","2"] :
            fillYorN = True  

        #    472 体验台桌下灯箱与产品保持同侧
        if testCol =="BV" and testValue[0:1] not in ["1","5","3"] and  imageType =="3" :
            fillYorN = True         
        # 4.7.3 机模陈列整齐
        if testCol =="BW" and testValue[0:1] not in ["1","3"]:
            fillYorN = True       
        #4.7.4 产品及陈列规范 BX 1
        #空置机位
        if testCol =="BY" and testValue[0:1] not in ["1","4"]:
            fillYorN = True     
        #4.5.1 所有手机真机 Retail Mode 程序100%自动运行
        if testCol =="BZ" and testValue[0:1] not in ["1","4"]:
            fillYorN = True  
        #4.5.2 价签自动跳转（Flod2 不考核） 4.5.3 电子价签（“价格信息”app）是否更新了最新的促销活动
        if testCol in ["CB","CA"] and testValue[0:1] not in ["1","5"]:
            fillYorN = True          
        #人员
        if testCol == "CC" and testValue[0:1] not in ["1","2"] :
            fillYorN = True 
        if testCol == "CD" and testValue[0:1] not in ["1","5"] :
            fillYorN = True 
        if testCol == "CE" and investorType=="SES" and testValue[0:1] not in ["1","3"] :
            fillYorN = True 
        if testCol == "CG" and otherValue(testCol,rowNumber,-1)[0:1] =="1":
            testp = otherValue(testCol,rowNumber,-60) if otherValue(testCol,rowNumber,-60)  else "无"
            YesOrNo ,name = namematch(testValue,testp)
            ws2["CG{}".format(rowNumber)].value = name
            if not YesOrNo:
                fillYorN = True 
        #知识考核  多选题 == -1 是无该选项 ， 多项选择判断条件成立则为 错误答案
        if datiren=="1":
            if testCol =="CH":#C    
                fillYorN = True if testValue[0:1]!="C" else False
            if testCol =="CI":#B
                fillYorN = True if testValue[0:1]!="B" else False

            if testCol =="CJ" :#ABCD
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True     

            if testCol =="CK":#ACD 
                if testValue.find("A、") == -1 or testValue.find("B、") != -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True    
            if testCol =="CL":#C    
                fillYorN = True if testValue[0:1]!="C" else False
            if testCol =="CM":#A   
                fillYorN = True if testValue[0:1]!="A" else False
            if testCol =="CN" :#AD
                if testValue.find("A、") == -1 or testValue.find("B、")!= -1 or testValue.find("C、")!= -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True     
            if testCol =="CO":#B
                fillYorN = True if testValue[0:1]!="B" else False
            if testCol =="CP" :#ABCD
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True     
            if testCol =="CQ":#B
                fillYorN = True if testValue[0:1]!="B" else False
        #竞品 CR 1
        
        # 违规  
        if testCol == "CT" and testValue[0:2]not in ["1.", "2.", "3."]:
            fillYorN = True
        if testCol == "CV" and testValue[0:2]not in ["7.", "8.", "9."]:
            fillYorN = True
        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            queDic["C{}".format(rowNumber)].append(
                "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
for r in range(2, ws2.max_row+1):
    valueIn = " ".join(queDic["C{}".format(r)])
    ws2["C{}".format(r)].value = valueIn

# 保存

print("saving......")
wb.save(fileName.replace(".xlsx","-修改.xlsx"))

wb.close()
