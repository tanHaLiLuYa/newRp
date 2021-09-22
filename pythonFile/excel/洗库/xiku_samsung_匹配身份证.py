import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter



def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a
def namematch(str1,str2):
    name1= "".join(re.findall(namepat,str1))
    list2= re.findall(namepat,str2)
    if name1 not in list2:
        return False , name1
    else :
        return True , name1



def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)



def otherValue(letter,r, num):
    tNum = column_index_from_string(letter)
    letternew= get_column_letter(tNum+num)
    return ws2["{}{}".format(letternew,r)].value



print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"C:\Users\tpp\Desktop\临时"


os.chdir(root_path)

fileName = "【W38】【洗库】0922_0830.xlsx"  

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = re.compile(r"[A-Z]{1,2}[0-9]{1,3}.{0,1}[0-9]{0,1}[.]")
matchPatternID =re.compile(r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}") 
namepat = re.compile(r"[\u4e00-\u9fa5]{1,12}")

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.active



queDic = {}
titleDic = {}

# 开始列 X,结束列 ,问题记录成字典内的列表
starColumnName = "W"
endColumnName = "BO"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
# for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
#     if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
#         titleDic["{}".format(get_column_letter(c))] = re.search(
#             matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    # queDic["C{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)
        print("列：", testCol, "行：", rowNumber)
        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"


        #答题人所在列
        # datiren = ws2["CB{}".format(rowNumber)].value if ws2["CB{}".format(rowNumber)].value else "none"
        # datiren =datiren[0:1]
        # # ————————————————————————————————————————————————————————————————————————————

        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        # 身份证
        if testCol == "X":
            testIDList = otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                # queDic["C{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                #                                        format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))
     
         
        # 答题人姓名 正确与否
        if testCol == "CN" and otherValue(testCol,rowNumber,-1)[0:1] =="1":
            testp = otherValue(testCol,rowNumber,-67) if otherValue(testCol,rowNumber,-67)  else "无"
            YesOrNo ,name = namematch(testValue,testp)
            ws2["CN{}".format(rowNumber)].value = name
            if not YesOrNo:
                fillYorN = True 
        #知识考核  多选题 == -1 是无该选项 ， 多项选择判断条件成立则为 错误答案
        # if datiren=="1":
        #     if testCol =="DD":
        #         if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") != -1 \
        #                 or testValue.find("D、") == -1:
        #             fillYorN = True
       
        #     if testCol =="DE":
        #         if testValue.find("A、") == -1 or testValue.find("B、") != -1 or testValue.find("C、") == -1 \
        #                 or testValue.find("D、") == -1:
        #             fillYorN = True
        #     if testCol =="DF":
        #         if testValue.find("A、") != -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
        #                 or testValue.find("D、") == -1:
        #             fillYorN = True

        #     if testCol =="DG":
        #         fillYorN = True if testValue.find("B、") == -1 else False 
        #     if testCol =="DH":
        #         fillYorN = True if testValue.find("B、") == -1 else False 
        #     if testCol =="DI":
        #         fillYorN = True if testValue.find("C、") == -1 else False 
        #     if testCol =="DJ":
        #         fillYorN = True if testValue.find("D、") == -1 else False 

        #     if testCol =="DK":
        #         if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") != -1 \
        #                 or testValue.find("D、") == -1:
        #             fillYorN = True

        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            # queDic["B{}".format(rowNumber)].append(
            #     "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
# for r in range(2, ws2.max_row+1):
#     valueIn = " ".join(queDic["B{}".format(r)])
#     ws2["B{}".format(r)].value = valueIn

# 保存修改

print("saving......")

wb.save(fileName.replace(".xlsx","-修改.xlsx"))

wb.close()
