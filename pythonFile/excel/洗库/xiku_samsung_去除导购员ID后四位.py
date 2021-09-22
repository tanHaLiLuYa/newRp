import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter



def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a
def namematch(str1,str2):
    name1= "".join(re.findall(namepat,str1))
    list2= re.findall(namepat,str2)
    if name1 not in list2:
        return False , name1
    else :
        return True , name1






print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"E:\work\tpp\samsung\2021年\汇总\4-7月已查导购员\file\output"


os.chdir(root_path)

fileName = "新建 Microsoft Excel 工作表.xlsx"  

matchPatternID =re.compile(r"[\u4e00-\u9fa5]{2,10}") 

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.active



for i  in range(2,ws2.max_row):
    testValue =ws2["S{}".format(i)].value 
    if not re.match(matchPatternID,  testValue):
        pass
    else:
        ws2["S{}".format(i)].value = ",".join(re.findall(matchPatternID, testValue))
#  


print("saving......")

wb.save(fileName.replace(".xlsx","-修改.xlsx"))

wb.close()
