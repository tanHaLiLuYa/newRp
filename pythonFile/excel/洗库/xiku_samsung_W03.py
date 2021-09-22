import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter



def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a
def namematch(str1,str2):
    name1= "".join(re.findall(namepat,str1))
    list2= re.findall(namepat,str2)
    if name1 not in list2:
        return False , name1
    else :
        return True , name1
def N20YANGJI(list1, list2):
    if len(set(list1)) != len(list1) or len(list2) % 2 != 0:
        return True
    else:
        listTest = []
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])

        for i in listTest:
            if i[0].isdigit():
                if int(i[0]) > 2 or (int(i[0]) > 1 and i[1] != "金"):
                    return True
            else:
                if int(i[1]) > 2 or (int(i[1]) > 1 and i[0] != "金"):
                    return True
        return False


def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)

        # listTest = []
        # for i in range(0,len(list2),2):
        #     listTest.append(list2[i]+list2[i+1])
        # for i in listTest:
        #     if i[0].isdigit():


def strToDicForTab(strIn):
    a = {}
    a["T720"] = strIn[strIn.find("T720")+4] if strIn.find("T720") != -1 else 0
    a["T510"] = strIn[strIn.find("T510")+4] if strIn.find("T510") != -1 else 0
    a["T860"] = strIn[strIn.find("T860")+4] if strIn.find("T860") != -1 else 0
    a["P610"] = strIn[strIn.find("P610")+4] if strIn.find("P610") != -1 else 0

    return a


def testTabFun(dic, dic_):
    for i in ["P610", "T860", "T510", "T720"]:
        if int(dic[i]) != int(dic_[i]):
            return True
    return False


def otherValue(letter,r, num):
    tNum = column_index_from_string(letter)
    letternew= get_column_letter(tNum+num)
    return ws2["{}{}".format(letternew,r)].value



print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"E:\work\tpp\samsung\2021年\W03\洗库"


os.chdir(root_path)

fileName = "【W03】【洗库】【有投资】0115_1430.xlsx"

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = re.compile(r"[A-W]{1,2}[0-9]{1,3}.{0,1}[0-9]{0,1}[.]")
matchPatternID =re.compile(r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}") 
namepat = re.compile(r"[\u4e00-\u9fa5]{2,12}")

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.copy_worksheet(wb["提交"])
ws2.title = "check1"


queDic = {}
titleDic = {}

# 开始列 X,结束列 ,问题记录成字典下的列表
starColumnName = "X"
endColumnName = "CV"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
    if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
        titleDic["{}".format(get_column_letter(c))] = re.search(
            matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    queDic["C{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)
        print("列：", testCol, "行：", rowNumber)
        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"

        investorType = ws2["K{}".format(rowNumber)].value
        investorType = investorType[0:3]  # 投资类型
        imageType = ws2["T{}".format(rowNumber)].value
        imageType = imageType[0:2]  # 新旧形象

        datiren = ws2["CH{}".format(rowNumber)].value if ws2["CH{}".format(rowNumber)].value else "none"
        datiren =datiren[0:1]
        # ————————————————————————————————————————————————————————————————————————————

        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        # 身份证
        if testCol == "X":
            testIDList = otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                queDic["C{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                                                       format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))

        # 1 门店形象 店外露出
        # ————————————————————————————————————————————————————————————————————————————

        if testCol in ["AB", "Z", "AA","AC","CR","BU"] and testValue[0:1] != "1" :  #, 
            fillYorN = True

        if testCol in ["AD"] and testValue[0:1] not in ["1", "4"] :
            fillYorN = True

        # 体验台销售柜
        if testCol in ["AG", "AJ", "AM", "AP"] :
            touru = otherValue(testCol,rowNumber, -1) if otherValue(testCol,rowNumber, -1) else 0
            fillYorN = True if int(touru) != int(testValue) else False
        #魔切卡片
        if testCol == "AQ" and otherValue(testCol,rowNumber,3)=="Y":
            fillYorN = True if testValue[0:1] != "1" else False
        if testCol =="AS" and otherValue(testCol,rowNumber,1)=="Y":
            fillYorN = True if testValue != otherValue(testCol,rowNumber,2) else False

        # #   W21 KV更换
        if testCol == "BD" and investorType=="SES" and testValue[0:1] not in ["1", "3"] :
            fillYorN = True

        # #   W21 紫紫物料
        if testCol == "BE" and investorType in ["SIS","SES"] and testValue[0:1] not in ["1", "3"] :
            fillYorN = True                     

        # W21 机模/真机
        if testCol == "BG" and otherValue(testCol,rowNumber,-1)=="Y" and investorType in ["SIS","SES"]:
            fillYorN = True if testValue[0:1] != "1" else False

        # Fold2 机模/真机
        if testCol == "BH" and otherValue(testCol,rowNumber,-6)=="Y":
            fillYorN = True if testValue[0:1] != "1" else False

        # S20 FE 多色
        if testCol== "BI" and imageType=="2." and otherValue(testCol,rowNumber,2)=="Y":
            tiyantainum = int(otherValue(testCol,rowNumber,3))
            if investorType == "SIS" and tiyantainum >= 3:
                fillYorN= True if testValue[0:1]  != "1" else False
            elif investorType == "SES":
                fillYorN= True if testValue[0:1]  != "1" else False
            else:
                fillYorN = True if testValue[0:1] not in ["1","2"] else False

        if testCol== "BJ" and imageType=="1." and otherValue(testCol,rowNumber,1)=="Y":
            if investorType in ["SCS"  ,"STS","ABS"] :
                fillYorN= True if testValue[0:1] not in ["1","2"] else False
            else:
                fillYorN = True if testValue[0:1] != "1" else False


        #新平台
        if testCol == "BM" and imageType=="1."  and investorType in ["SES","SIS"]  :
            fillYorN = True if testValue[0:1] not in ["1", "4"]  else False
        if testCol == "BN" and imageType=="1." and otherValue(testCol,rowNumber,-12)=="Y" and investorType in ["SES","SIS"]  :
            fillYorN = True if testValue[0:1] not in ["1", "4"]  else False

        #   W21 5G使用托架陈列
        if testCol== "BO" and investorType in ["SES","SIS"] and otherValue(testCol,rowNumber,-9)=="Y":
            fillYorN= True if testValue[0:1]!="1" else False

        #  W21 5G托架使用规范   
        if testCol== "BP"  and investorType in ["SES","SIS"] and otherValue(testCol,rowNumber,-10)=="Y":
            fillYorN= True if testValue[0:1]!="1" else False

        #4.9 传统形象门店S20 FE 陈列套装
        if testCol =="BQ"  and imageType=="2." and otherValue(testCol,rowNumber,-6)=="Y":
            fillYorN= True if testValue[0:1]!="1" else False
        #4.10 新形象门店S20 FE 3机托架陈列
        if testCol =="BR"  and imageType=="1." and otherValue(testCol,rowNumber,-7)=="Y":
            fillYorN= True if testValue[0:1] not in ["1", "7"]   else False   

        #4.14 体验台桌下灯箱与产品保持同侧
        if testCol =="BS" and imageType=="2." and investorType in ["SES","SIS"]:
            fillYorN= True if testValue[0:1] not in ["1","5","3"] else False    
        #retail
        if testCol == "CA" :
            fillYorN= True if testValue[0:1] not in ["1","4"] else False
        if testCol == "CB" :
            fillYorN= True if testValue[0:1] not in ["1","5"] else False
    
        
        #a机模陈列
        if testCol == "BT" :
            fillYorN= True if testValue[0:1] not in ["1","3"] else False
        #   4.16 产品陈列规范 BU 1
        #   4.17 机位空置 	
        if testCol == "BV" :
            fillYorN= True if testValue[0:1] not in ["1","4"] else False

        # 保护壳
        if testCol == "BW" and investorType =="SES":
            fillYorN= True if testValue[0:1] not in ["1"] else False
        # Buds Live真机露出
        if testCol == "BX" and investorType =="SES":
            fillYorN= True if testValue[0:1] not in ["1","2","4"] else False

        #配件墙陈列
        if testCol =="BY" and imageType=="2." and investorType =="SES":
            fillYorN= True if testValue[0:1] not in ["1","3"] else False

        if testCol =="BZ" and imageType=="1." and investorType =="SES":
            fillYorN= True if testValue[0:1] not in ["1","4","5"] else False         
        
        #人员
        if testCol == "CC" and testValue[0:1] not in ["1","2"] :
            fillYorN = True 
        if testCol == "CD" and testValue[0:1] not in ["1","5","6"] :
            fillYorN = True 
        if testCol == "CE" and investorType=="SES" and testValue[0:1] not in ["1","3"] :
            fillYorN = True 
        if testCol == "CI" and otherValue(testCol,rowNumber,-1)[0:1] =="1":
            YesOrNo ,name = namematch(testValue,otherValue(testCol,rowNumber,-62))
            ws2["CI{}".format(rowNumber)].value = name
            if not YesOrNo:
                fillYorN = True 

                
                
        #知识考核
        if datiren=="1":
            if testCol =="CJ":#D    
                fillYorN = True if testValue[0:1]!="D" else False
            if testCol =="CK":#B 
                fillYorN = True if testValue[0:1]!="B" else False
            if testCol =="CL":#C
                fillYorN = True if testValue[0:1]!="C" else False

            if testCol =="CM" :#ABCD
                if testValue.find("A、") != -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True     

            if testCol =="CN" :#B
                fillYorN = True if testValue[0:1]!="B" else False
            if testCol =="CO":#B  
                fillYorN = True if testValue[0:1]!="B" else False
            if testCol =="CP" :#D
                fillYorN = True if testValue[0:1]!="D" else False
            if testCol =="CQ" :#A D
                if testValue.find("A、") == -1 or testValue.find("B、") != -1 or testValue.find("C、") != -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True     

        #竞品 Cr 1
        
        # 违规  
        if testCol == "CT" and testValue[0:2]not in ["1.", "2.", "3."]:
            fillYorN = True
        if testCol == "CV" and testValue[0:2]not in ["7.", "8.", "9."]:
            fillYorN = True
        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            queDic["C{}".format(rowNumber)].append(
                "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
for r in range(2, ws2.max_row+1):
    valueIn = " ".join(queDic["C{}".format(r)])
    ws2["C{}".format(r)].value = valueIn

# 保存

print("saving......")
wb.save(fileName.replace(".xlsx","-修改.xlsx"))

wb.close()
