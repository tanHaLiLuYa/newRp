import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter



def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a
def namematch(str1,str2):
    name1= "".join(re.findall(namepat,str1))
    list2= re.findall(namepat,str2)
    if name1 not in list2:
        return False , name1
    else :
        return True , name1
def N20YANGJI(list1, list2):
    if len(set(list1)) != len(list1) or len(list2) % 2 != 0:
        return True
    else:
        listTest = []
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])

        for i in listTest:
            if i[0].isdigit():
                if int(i[0]) > 2 or (int(i[0]) > 1 and i[1] != "金"):
                    return True
            else:
                if int(i[1]) > 2 or (int(i[1]) > 1 and i[0] != "金"):
                    return True
        return False


def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)

        # listTest = []
        # for i in range(0,len(list2),2):
        #     listTest.append(list2[i]+list2[i+1])
        # for i in listTest:
        #     if i[0].isdigit():


def strToDicForTab(strIn):
    a = {}
    a["T720"] = strIn[strIn.find("T720")+4] if strIn.find("T720") != -1 else 0
    a["T510"] = strIn[strIn.find("T510")+4] if strIn.find("T510") != -1 else 0
    a["T860"] = strIn[strIn.find("T860")+4] if strIn.find("T860") != -1 else 0
    a["P610"] = strIn[strIn.find("P610")+4] if strIn.find("P610") != -1 else 0

    return a


def testTabFun(dic, dic_):
    for i in ["P610", "T860", "T510", "T720"]:
        if int(dic[i]) != int(dic_[i]):
            return True
    return False


def otherValue(letter,r, num):
    tNum = column_index_from_string(letter)
    letternew= get_column_letter(tNum+num)
    return ws2["{}{}".format(letternew,r)].value



print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"E:\work\tpp\samsung\2021年\04月\W16\洗库"


os.chdir(root_path)

fileName = "【W16】【洗库】0419_1000.xlsx"  

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = re.compile(r"[A-Z]{1,2}[0-9]{1,3}.{0,1}[0-9]{0,1}[.]")
matchPatternID =re.compile(r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}") 
namepat = re.compile(r"[\u4e00-\u9fa5]{1,12}")

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.copy_worksheet(wb["提交"])
ws2.title = "check1"


queDic = {}
titleDic = {}

# 开始列 X,结束列 ,问题记录成字典内的列表
starColumnName = "X"
endColumnName = "CZ"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
    if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
        titleDic["{}".format(get_column_letter(c))] = re.search(
            matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    queDic["C{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)
        print("列：", testCol, "行：", rowNumber)
        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"

        investorType = ws2["K{}".format(rowNumber)].value
        investorType = investorType[0:3]  # 投资类型
        imageType = ws2["T{}".format(rowNumber)].value
        imageType = imageType[0:1]  # 新旧形象

        datiren = ws2["CN{}".format(rowNumber)].value if ws2["CN{}".format(rowNumber)].value else "none"
        datiren =datiren[0:1]
        # ————————————————————————————————————————————————————————————————————————————

        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        # 身份证
        if testCol == "X":
            testIDList = otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                queDic["C{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                                                       format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))

        # 1 门店形象 店外露出
        # ————————————————————————————————————————————————————————————————————————————

        if testCol in ["Z", "AA","AC","AB","BQ","BY","BZ","CA","BR","BS","CC" ,"CD","CW"] and testValue[0:1] != "1" :  #, 
            fillYorN = True

        if testCol in ["AD"] and testValue[0:1] not in ["1", "4"] :
            fillYorN = True

        # 体验台销售柜
        if testCol in ["AJ", "AG", "AM", "AP"] :
            touru = otherValue(testCol,rowNumber, -1) if otherValue(testCol,rowNumber, -1) else 0
            fillYorN = True if int(touru) != int(testValue) else False
        # # S21 机模陈列
        # if testCol =="AV" and otherValue(testCol,rowNumber, 1)=="Y" and testValue[0:1] not in ["1"] :
        #     fillYorN = True

        #4.1.2 A52多色出样
        if testCol =="AW" and otherValue(testCol,rowNumber, -2)=="Y" and otherValue(testCol,rowNumber, 1)>0 and testValue[0:1] not in ["4"] :
            if imageType =="3" and otherValue(testCol,rowNumber, 1)>3:
                fillYorN = True
            if imageType =="3" and otherValue(testCol,rowNumber, 1)<=3:
                fillYorN = True if  testValue[0:1] =="5" else False
            if imageType!= "3" and otherValue(testCol,rowNumber, 1)>2:
                fillYorN = True
            if imageType!="3" and  otherValue(testCol,rowNumber, 1)<=2:
                fillYorN = True if  testValue[0:1] in ["5","1"] else False


        #4.2.2 LFD（店面入口静态画面）是否有播放Z系列 KV画面内容	4.2.3 LFD（电视屏幕动态播放）是否有播放A52产品KV画面内容	

        if testCol in ["AY","AZ"] and testValue[0:1] not in ["1", "3"] :
            fillYorN = True

        # 4.4.2.2 S21 陈列套装组
        if testCol =="BA" and otherValue(testCol,rowNumber, -8)=="Y" and imageType in ["1"] :
            fillYorN = True if testValue[0:1] not in ["1", "5"] else False

        #S21
        if testCol =="BC" and otherValue(testCol,rowNumber, -1)=="Y" and imageType in ["1","2"] and testValue[0:1] not in ["1"] :
            fillYorN = True
        if testCol =="BD" and otherValue(testCol,rowNumber, -2)=="Y" and imageType in ["1","2"] and testValue[0:1] not in ["1"] :
            fillYorN = True
        if testCol =="BE" and otherValue(testCol,rowNumber, -3)=="Y" and imageType in ["3"] and testValue[0:1] not in ["1"] :
            fillYorN = True
        if testCol =="BF" and otherValue(testCol,rowNumber, -4)=="Y" and imageType in ["3"] and testValue[0:1] not in ["1"] :
            fillYorN = True
        if testCol =="BH" and otherValue(testCol,rowNumber, -1)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True
        if testCol =="BI" and otherValue(testCol,rowNumber, -2)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True
        if testCol =="BK" and otherValue(testCol,rowNumber, -1)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True
        if testCol =="BL" and otherValue(testCol,rowNumber, -2)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True
        #A52
        if testCol =="BM" and otherValue(testCol,rowNumber, -18)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True
        if testCol =="BN" and otherValue(testCol,rowNumber, -19)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True        

        #4.1.2 A52多色出样
        if testCol =="AW" and otherValue(testCol,rowNumber, 1)=="Y" and otherValue(testCol,rowNumber, 2)>0 and testValue[0:1] not in ["4"] :
            if  imageType =="3" and otherValue(testCol,rowNumber, 2)>3:
                fillYorN = True
            if imageType =="3" and otherValue(testCol,rowNumber, 2)<=3:
                fillYorN = True if  testValue[0:1] =="5" else False
            if imageType!= "3" and otherValue(testCol,rowNumber, 2)>2:
                fillYorN = True
            if imageType!="3" and  otherValue(testCol,rowNumber, 2)<=2:
                fillYorN = True if  testValue[0:1] in ["5","1"] else False
        
        #促销物料
        if testCol =="BO" and otherValue(testCol,rowNumber, -21)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True     
        if testCol =="BP" and otherValue(testCol,rowNumber, -23)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True  
        
        # 4.4.3 体验台桌面降价贴露出	 BQ 1
        #4.4.4 春日钜惠促销KV物料露出	BR 1 BS1
        #A52促销
        if testCol =="BT" and otherValue(testCol,rowNumber, -25)=="Y" and testValue[0:1] not in ["1"] :
            fillYorN = True  
        #配件桌
        if testCol in ["BU"] and testValue[0:1] not in ["1", "3","4"] :
            fillYorN = True
        #产品强
        if testCol in ["BV"] and testValue[0:1] not in ["1", "3","5","6"] and imageType !="3":
            fillYorN = True
        if testCol in ["BW"] and testValue[0:1] not in ["1","3","5", "8","9"] and imageType !="3":
            fillYorN = True
        if testCol in ["BX"] and testValue[0:1] not in ["1","3","5", "6"] and imageType =="3":
            fillYorN = True
        #体验内容 By BZ CA 1
        #基础规范
        if testCol in ["CB"] and testValue[0:1] not in ["1", "5","3"] and imageType =="3":
            fillYorN = True
        # CC CD 1
        if testCol in ["CE"] and testValue[0:1] not in ["1","4"] :
            fillYorN = True

        #操作题目
        if testCol in ["CF","CH","CJ"]  and testValue[0:1] not in ["4","1"] :
            fillYorN = True  
        if testCol =="CI"  and testValue[0:1] not in ["5","1"] :
            fillYorN = True  
        if testCol =="CG"  and testValue[0:1] not in ["6","1"] :
            fillYorN = True  

        #人员
        if testCol == "CK" and testValue[0:1] not in ["1","2"] :
            fillYorN = True 
        if testCol == "CL" and testValue[0:1] not in ["1","5","6"] :
            fillYorN = True 
        if testCol == "CM" and testValue[0:1] not in ["1","3"] and investorType=="SES":
            fillYorN = True    

         
        #答题人姓名 正确与否
        if testCol == "CO" and otherValue(testCol,rowNumber,-1)[0:1] =="1":
            testp = otherValue(testCol,rowNumber,-68) if otherValue(testCol,rowNumber,-68)  else "无"
            YesOrNo ,name = namematch(testValue,testp)
            ws2["CO{}".format(rowNumber)].value = name
            if not YesOrNo:
                fillYorN = True 
        #知识考核  多选题 == -1 是无该选项 ， 多项选择判断条件成立则为 错误答案
        if datiren=="1":
            if testCol =="CP":
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") != -1:
                    fillYorN = True
            if testCol =="CQ":
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") != -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True
            if testCol =="CR":
                fillYorN = True if testValue.find("120Hz高刷新率") == -1 else False 
            if testCol =="CS":
                fillYorN = True if testValue.find("后置单独的微距镜头") == -1 else False 
            if testCol =="CT":
                if testValue.find("A、") != -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:                   
                    fillYorN = True

        #竞品 CW 1
        
        # 违规  
        # if testCol == "CC" and testValue[0:2]not in ["1.", "2.", "3."]:
        #     fillYorN = True
        if testCol == "CY" and testValue[0:2]not in ["7.", "8.", "9."]:
            fillYorN = True
        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            queDic["C{}".format(rowNumber)].append(
                "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
for r in range(2, ws2.max_row+1):
    valueIn = " ".join(queDic["C{}".format(r)])
    ws2["C{}".format(r)].value = valueIn

# 保存

print("saving......")

wb.save(fileName.replace(".xlsx","-修改.xlsx"))

wb.close()
