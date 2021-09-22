import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from tqdm import tqdm


def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a


def N20YANGJI(list1, list2):
    if len(set(list1)) != len(list1) or len(list2) % 2 != 0:
        return True
    else:
        listTest = []
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])

        for i in listTest:
            if i[0].isdigit():
                if int(i[0]) > 2 or (int(i[0]) > 1 and i[1] != "金"):
                    return True
            else:
                if int(i[1]) > 2 or (int(i[1]) > 1 and i[0] != "金"):
                    return True
        return False


def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)

        # listTest = []
        # for i in range(0,len(list2),2):
        #     listTest.append(list2[i]+list2[i+1])
        # for i in listTest:
        #     if i[0].isdigit():


def strToDicForTab(strIn):
    a = {}
    a["T720"] = strIn[strIn.find("T720")+4] if strIn.find("T720") != -1 else 0
    a["T510"] = strIn[strIn.find("T510")+4] if strIn.find("T510") != -1 else 0
    a["T860"] = strIn[strIn.find("T860")+4] if strIn.find("T860") != -1 else 0
    a["P610"] = strIn[strIn.find("P610")+4] if strIn.find("P610") != -1 else 0

    return a


def testTabFun(dic, dic_):
    for i in ["P610", "T860", "T510", "T720"]:
        if int(dic[i]) != int(dic_[i]):
            return True
    return False


print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"D:\work\samsung\W47周报\洗库"

os.chdir(root_path)

fileName = "W46洗库_1109_1150.xlsx"

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = r"[A-W]{1,2}[1-9]{1,3}.{0,1}[1-9]{0,1}[.]"
matchPatternID = r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}"

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.copy_worksheet(wb["提交"])
ws2.title = "check1"


queDic = {} 
titleDic = {}

# 开始列 X,结束列BL,问题记录成字典下的列表
starColumnName = "X"
endColumnName = "BV"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
    if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
        titleDic["{}".format(get_column_letter(c))] = re.search(
            matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    queDic["C{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)
        print("列：", testCol, "行：", rowNumber)
        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"

        investorType = ws2["K{}".format(rowNumber)].value
        investorType = investorType[0:3]  # 投资类型
        imageType = ws2["T{}".format(rowNumber)].value
        imageType = imageType[0:2]  # 新旧形象

        # ————————————————————————————————————————————————————————————————————————————
        testTab = ws2["AF{}".format(rowNumber)].value if ws2["AF{}".format(
            rowNumber)].value else "error"  # Tab
        testIDList = ws2["Y{}".format(rowNumber)].value.split(
            ",") if ws2["Y{}".format(rowNumber)].value else "error"  # 导购员姓名IDlist
        testAnswerP = ws2["BF{}".format(rowNumber)].value if ws2["BF{}".format(
            rowNumber)].value else "error" # 答题人
        # ————————————————————————————————————————————————————————————————————————————
        FOLDShop = ws2["AH{}".format(rowNumber)].value  # FOLD
        FE = ws2["AL{}".format(rowNumber)].value  # FE
        TIYANTAINUM = ws2["AO{}".format(rowNumber)].value  # 体验台数量
        SHOPCat = ws2["BT{}".format(rowNumber)].value#SES店面类型
        PHONEplus = ws2["BU{}".format(rowNumber)].value# phone+ 
        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        # 身份证
        if testCol == "X":
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                queDic["C{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                                                       format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))

        # 1 门店形象 店外露出
        # ————————————————————————————————————————————————————————————————————————————
        if testCol in ["Z", "AA", "AB", "AC", "BN"] and testValue[0:2] != "1.":  # 
            fillYorN = True
        if testCol == "AD" and testValue[0:2] not in ["1.", "4."]:
            fillYorN = True
        # # TAB
        # if testTab != "P610 0":  # tab需要有投入覆盖 才匹配
        #     if testCol == "AE":
        #         if re.findall(matchPatternTab, testValue):
        #             # print("{}{}".format(get_column_letter(colNumber),rowNumber),re.findall(matchPatternColor, testValue))
        #             testValue = "," .join(
        #                 re.findall(matchPatternTab, testValue))
        #             ws2["{}{}".format(get_column_letter(colNumber),
        #                               rowNumber)].value = testValue
        #         else:
        #             testValue = "error"
        #         dictab = strToDicForTab(testValue)
        #         dictab_tab = strToDicForTab(testTab)
        #         # print(dictab, dictab_tab, rowNumber)
        #         if testTabFun(dictab, dictab_tab):
        #             fillYorN = True
        # 3 陈列体验
        # ————————————————————————————————————————————————————————————————————————————
        # 服务卡
        if testCol == "AJ"  and testValue[0:2] not in ["1."]:
            fillYorN = True
        # 立牌
        if testCol == "AK"  and testValue[0:2] not in ["1."]:
            fillYorN = True
        # LFD播放画面
        if testCol == "AM" and testValue[0:2] not in ["1.", "3."]:
            fillYorN = True

        # Fold2 机模真机露出
        if testCol == "AN" and FOLDShop == "Y" and testValue[0:2] not in ["1."]:
            fillYorN = True

        # Fe 多色出样
        if testCol == "AP" and FE == "Y":
            fillYorN = True if testValue[0:2] not in [
                    "1.", "2."] else False
        # 新品台陈列
        if testCol == "AQ" and imageType == "1.":
            fillYorN = True if testValue[0:2] not in ["1.", "4."] else False
        # 异形真机露出??????
        if testCol == "AR":
            if imageType != "1." and SHOPCat not in ["A", "B"]:
                pass
            else:
                fillYorN = True if testValue[0:2] not in [
                    "1.", "6."] else False
        # 非异形 陈列
        if testCol == "AS" and testValue[0:2] not in ["1.", "6.", "4."]:
            fillYorN = True
        # FE tongce
        if testCol == "AT" and imageType == "2." and testValue[0:2] not in ["1.", "3.", "4."]:
            fillYorN = True
        # FE 套餐陈列
        if testCol == "AU" and imageType == "2.":
            fillYorN = True if testValue[0:2] != "1." else False
        # S20FE 托架
        if testCol == "AV" and imageType == "1.":
            fillYorN = True if testValue[0:2] != "1." else False
        # 同侧陈列
        if testCol == "AW" and imageType == "2.":
            fillYorN = True if testValue[0:2] not in["5.",
                                                     "3.", "1."] else False
        # 保护壳 露出
        if testCol == "AX" :
            fillYorN = True if testValue[0:2] != "1." else False
        # 保护壳 托架陈列
        if testCol == "AY"  and imageType == "2.":
            fillYorN = True if testValue[0:2] != "1." else False
        # Buds 样机
        if testCol == "AZ" :
            fillYorN = True if testValue[0:2] not in["4.",
                                                     "2.", "1."] else False


        # retail 播放
        if testCol == "BA" and testValue[0:2] not in ["1.", "4."]:
            fillYorN = True
        # retail 自动跳转到price label & jiagegenxin
        if testCol in ["BB", "BC"] and testValue[0:2] not in ["1.", "5."]:
            fillYorN = True
 
         # 违规陈列
        if testCol in ["BD"] :
            fillYorN = True if testValue[0:2] != "1." else False

        # 人员部分
        # ————————————————————————————————————————————————————————————————————————————
        # 着装
        if testCol == "BE":
            fillYorN = True if testValue[0:2] != "1." else False
        # 答题部分
        if testAnswerP[0:2] == "1.":
            if testCol in ["BH"]:
                if testValue.find("A、") == -1 or testValue.find("B、") != -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True
            if testCol in ["BI"]:
                if testValue.find("A、") != -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True
            if testCol in ["BJ"] and testValue[0:2] != "B、":
                fillYorN = True
            if testCol in ["BK", "BL"] and testValue[0:2] != "D、":
                fillYorN = True
            #
            if PHONEplus == "Y":
                if testCol in ["BM"] and testValue[0:2] != "B、":
                    fillYorN = True



        # 违规

        if testCol == "BP" and testValue[0:2]not in ["1.", "2.", "3."]:
            fillYorN = True
        if testCol == "BR" and testValue[0:2]not in ["7.", "8.", "9."]:
            fillYorN = True

        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            queDic["C{}".format(rowNumber)].append(
                "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
for r in tqdm(range(2, ws2.max_row+1)):
    valueIn = "".join(queDic["C{}".format(r)])
    ws2["C{}".format(r)].value = valueIn

# 保存
fileNameNew = "check.xlsx"
print("saving......")
wb.save(fileNameNew)

wb.close()
