import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter



def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a


def N20YANGJI(list1, list2):
    if len(set(list1)) != len(list1) or len(list2) % 2 != 0:
        return True
    else:
        listTest = []
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])

        for i in listTest:
            if i[0].isdigit():
                if int(i[0]) > 2 or (int(i[0]) > 1 and i[1] != "金"):
                    return True
            else:
                if int(i[1]) > 2 or (int(i[1]) > 1 and i[0] != "金"):
                    return True
        return False


def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)

        # listTest = []
        # for i in range(0,len(list2),2):
        #     listTest.append(list2[i]+list2[i+1])
        # for i in listTest:
        #     if i[0].isdigit():


def strToDicForTab(strIn):
    a = {}
    a["T720"] = strIn[strIn.find("T720")+4] if strIn.find("T720") != -1 else 0
    a["T510"] = strIn[strIn.find("T510")+4] if strIn.find("T510") != -1 else 0
    a["T860"] = strIn[strIn.find("T860")+4] if strIn.find("T860") != -1 else 0
    a["P610"] = strIn[strIn.find("P610")+4] if strIn.find("P610") != -1 else 0

    return a


def testTabFun(dic, dic_):
    for i in ["P610", "T860", "T510", "T720"]:
        if int(dic[i]) != int(dic_[i]):
            return True
    return False


def otherValue(letter,r, num):
    tNum = column_index_from_string(letter)
    letternew= get_column_letter(tNum+num)
    return ws2["{}{}".format(letternew,r)].value

def namematch(str1,str2):
    name1= "".join(re.findall(namepat,str1))
    list2= re.findall(namepat,str2)
    if name1 not in list2:
        return False , name1
    else :
        return True , name1
print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"E:\work\tpp\samsung\2021年\02月\W08\洗库"

os.chdir(root_path)

fileName = "【W09】【洗库】【无投资】0225_1530.xlsx"

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = r"[A-W]{1,2}[0-9]{1,3}.{0,1}[0-9]{0,1}[.]"
matchPatternID = r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}"
namepat = re.compile(r"[\u4e00-\u9fa5]{1,12}")

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.copy_worksheet(wb["提交"])
ws2.title = "check1"


queDic = {}
titleDic = {}

# 开始列 X,结束列 ,问题记录成字典下的列表
starColumnName = "U"
endColumnName = "BE"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
    if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
        titleDic["{}".format(get_column_letter(c))] = re.search(
            matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    queDic["B{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)

        print("列：", testCol, "行：", rowNumber)

        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"


        datiren = ws2["AT{}".format(rowNumber)].value if ws2["AT{}".format(rowNumber)].value else "none"
        datiren =datiren[0:1]
        # ————————————————————————————————————————————————————————————————————————————

        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        if testCol == "U":
            testIDList = otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                queDic["B{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                                                       format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))
        # 1 B 题目
        # ————————————————————————————————————————————————————————————————————————————

        if testCol in ["W","Y", "Z", "AA","X"] and testValue[0:1] != "1" :  #,
            fillYorN = True

        if testCol in ["AD","AB"] and testValue[0:1] !="2" :
            fillYorN = True
        if testCol == "AF" and testValue[0:1] not in ["1","3"]:
            fillYorN = True
        if testCol == "AG" and testValue[0:1] not in ["1","4"]:
            fillYorN = True
        if testCol == "AK" and otherValue(testCol,rowNumber,-3) == "N" and otherValue(testCol,rowNumber,-2)=="Y": 
            fillYorN = True if testValue[0:1] != "1" else False
        if testCol == "AL" and otherValue(testCol,rowNumber,-4) == "N" and otherValue(testCol,rowNumber,-2)=="Y": 
            fillYorN = True if testValue[0:1] != "1" else False
        if testCol == "AM" and otherValue(testCol,rowNumber,-5) == "Y" and otherValue(testCol,rowNumber,-4)=="Y": 
            fillYorN = True if testValue[0:1] != "1" else False
        if testCol == "AQ" and otherValue(testCol,rowNumber,-9) == "Y" and otherValue(testCol,rowNumber,-1)=="Y": 
            fillYorN = True if testValue[0:1] not in ["1","7"] else False
        if testCol == "AN" and otherValue(testCol,rowNumber,-6) == "Y" : 
            fillYorN = True if testValue[0:1] not in ["1","3"] else False
        if testCol == "AO" and otherValue(testCol,rowNumber,-7) == "Y" : 
            fillYorN = True if testValue[0:1] not in ["1","3"] else False

        #人员
        if testCol == "AR" and testValue[0:1] not in ["1", "2"] :
            fillYorN = True 
        if testCol == "AS" and testValue[0:1] not in ["1", "5","6"] and datiren=="1":
            fillYorN = True 
        if testCol == "AU" and otherValue(testCol,rowNumber,-1)[0:1] =="1":
            testp = otherValue(testCol,rowNumber,-25) if otherValue(testCol,rowNumber,-25)  else "无"
            YesOrNo ,name = namematch(testValue,testp)
            ws2["AU{}".format(rowNumber)].value = name
            if not YesOrNo:
                fillYorN = True 

        #知识考核
        if datiren=="1":
            if testCol =="AV":#A    
                fillYorN = True if testValue[0:1]!="A" else False
            if testCol =="AW":#A   
                fillYorN = True if testValue[0:1]!="A" else False
            if testCol =="AX":#D   
                fillYorN = True if testValue[0:1]!="D" else False
                
            if testCol =="AY" :#ABCD
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True     

            if testCol =="AZ" :#AC
                fillYorN = True if testValue[0:1]!="C" else False
            if testCol =="BA":#A   
                fillYorN = True if testValue[0:1]!="A" else False
            if testCol =="BB" :#CD
                if testValue.find("A、") != -1 or testValue.find("B、") != -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True     
            if testCol =="BC" :#ABCD
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True     
           

        
        # 违规  

        if testCol == "BH" and testValue[0:1]not in ["1", "2","3"]:
            fillYorN = True
        if testCol == "BJ" and testValue[0:1]not in ["7", "8", "9"]:
            fillYorN = True
        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            queDic["B{}".format(rowNumber)].append(
                "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
for r in range(2, ws2.max_row+1):
    valueIn = " ".join(queDic["B{}".format(r)])
    ws2["B{}".format(r)].value = valueIn

# 保存
fileNameNew = "check.xlsx"

print("saving......")
wb.save(fileNameNew)

wb.close()
