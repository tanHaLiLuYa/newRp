import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter



def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a


def N20YANGJI(list1, list2):
    if len(set(list1)) != len(list1) or len(list2) % 2 != 0:
        return True
    else:
        listTest = []
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])

        for i in listTest:
            if i[0].isdigit():
                if int(i[0]) > 2 or (int(i[0]) > 1 and i[1] != "金"):
                    return True
            else:
                if int(i[1]) > 2 or (int(i[1]) > 1 and i[0] != "金"):
                    return True
        return False


def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)

        # listTest = []
        # for i in range(0,len(list2),2):
        #     listTest.append(list2[i]+list2[i+1])
        # for i in listTest:
        #     if i[0].isdigit():


def strToDicForTab(strIn):
    a = {}
    a["T720"] = strIn[strIn.find("T720")+4] if strIn.find("T720") != -1 else 0
    a["T510"] = strIn[strIn.find("T510")+4] if strIn.find("T510") != -1 else 0
    a["T860"] = strIn[strIn.find("T860")+4] if strIn.find("T860") != -1 else 0
    a["P610"] = strIn[strIn.find("P610")+4] if strIn.find("P610") != -1 else 0

    return a


def testTabFun(dic, dic_):
    for i in ["P610", "T860", "T510", "T720"]:
        if int(dic[i]) != int(dic_[i]):
            return True
    return False


def otherValue(letter,r, num):
    tNum = column_index_from_string(letter)
    letternew= get_column_letter(tNum+num)
    return ws2["{}{}".format(letternew,r)].value


print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"E:\work\tpp\samsung\W51周报\洗库"

os.chdir(root_path)

fileName = "W51洗库_1215_1120.xlsx"

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = r"[A-W]{1,2}[0-9]{1,3}.{0,1}[0-9]{0,1}[.]"
matchPatternID = r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}"

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.copy_worksheet(wb["提交"])
ws2.title = "check1"


queDic = {}
titleDic = {}

# 开始列 X,结束列 ,问题记录成字典下的列表
starColumnName = "X"
endColumnName = "CP"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
    if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
        titleDic["{}".format(get_column_letter(c))] = re.search(
            matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    queDic["C{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)
        print("列：", testCol, "行：", rowNumber)
        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"

        investorType = ws2["K{}".format(rowNumber)].value
        investorType = investorType[0:3]  # 投资类型
        imageType = ws2["T{}".format(rowNumber)].value
        imageType = imageType[0:2]  # 新旧形象
        datiren = ws2["BX{}".format(rowNumber)].value if ws2["BX{}".format(rowNumber)].value else "none"
        datiren =datiren[0:1]
        # ————————————————————————————————————————————————————————————————————————————

        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        # 身份证
        if testCol == "X":
            testIDList = otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                queDic["C{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                                                       format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))

        # 1 门店形象 店外露出
        # ————————————————————————————————————————————————————————————————————————————

        if testCol in ["AB", "Z", "AA","AC","AX","AY","AZ","BR","CK"] and testValue[0:1] != "1" :  #,"AV","BN","BO","BZ"
            fillYorN = True

        if testCol in ["AD"] and testValue[0:1] not in ["1", "4"] :
            fillYorN = True

        # 体验台销售柜
        if testCol in ["AG", "AJ", "AM", "AP"] :
            touru = otherValue(testCol,rowNumber, -1) if otherValue(testCol,rowNumber, -1) else 0
            fillYorN = True if int(touru) != int(testValue) else False
        # W21 心系天下 灯箱片
        if testCol == "AW" and testValue[0:1] not in ["1", "3"] :
            fillYorN = True
                      
        # W21 心系天下 AX 1
        # 服务手册 AY 1
        # 体验台上露出双十二促销A4立牌 AZ 1
        
        # W21 KV画面内容 双十二促销KV
        if testCol in ["BB","BA"] and testValue[0:1] not in ["1", "3"]  :
            fillYorN = True
        # W21 机模/真机
        if testCol == "BC" and otherValue(testCol,rowNumber,1)=="Y":
            fillYorN = True if testValue[0:1] != "1" else False

        # Fold2 机模/真机
        if testCol == "BE" and otherValue(testCol,rowNumber,-10)=="Y":
            fillYorN = True if testValue[0:1] != "1" else False

        # S20 FE 多色
        if testCol== "BF" and imageType=="2." and otherValue(testCol,rowNumber,2)=="Y":
            tiyantainum = int(otherValue(testCol,rowNumber,3))
            if investorType == "SIS" and tiyantainum>=3:
                fillYorN= True if testValue[0:1]  != "1" else False
            else:
                fillYorN = True if testValue[0:1] not in ["1","2"] else False

        if testCol== "BG" and imageType=="1." and otherValue(testCol,rowNumber,1)=="Y":
            if investorType=="SCS" :
                fillYorN= True if testValue[0:1] not in ["1","2"] else False
            else:
                fillYorN = True if testValue[0:1] != "1" else False


        #新平台
        if testCol in ["BJ","BK"] and imageType=="1." and testValue[0:1] not in ["1", "4"]  :
            fillYorN = True

        #   W21 5G使用托架陈列
        if testCol== "BL" and otherValue(testCol,rowNumber,-8)=="Y":
            fillYorN= True if testValue[0:1]!="1" else False

        #  W21 5G托架使用规范
        if testCol== "BM" and otherValue(testCol,rowNumber,-9)=="Y":
            fillYorN= True if testValue[0:1]!="1" else False

        #4.9 传统形象门店S20 FE 陈列套装
        if testCol =="BN"  and imageType=="2." and otherValue(testCol,rowNumber,-6)=="Y":
            fillYorN= True if testValue[0:1]!="1" else False
        #4.10 新形象门店S20 FE 3机托架陈列
        if testCol =="BO"  and imageType=="1." and otherValue(testCol,rowNumber,-7)=="Y":
            fillYorN= True if testValue[0:1] !="1" else False   

        #4.14 体验台桌下灯箱与产品保持同侧
        if testCol =="BP" and imageType=="2.":
            fillYorN= True if testValue[0:1] not in ["1","5","3"] else False    
        #retail
        if testCol == "BT" :
            fillYorN= True if testValue[0:1] not in ["1","4"] else False
        if testCol == "BU" :
            fillYorN= True if testValue[0:1] not in ["1","5"] else False
        if testCol == "BV" :
            fillYorN= True if testValue[0:1] not in ["1","5"] else False
        
        #a机模陈列
        if testCol == "BQ" :
            fillYorN= True if testValue[0:1] not in ["1","3"] else False
        #   4.16 产品陈列规范 BR	1
        #   4.17 机位空置 	
        if testCol == "BS" :
            fillYorN= True if testValue[0:1] not in ["1","4"] else False

     
        # # Buds Live真机露出
        # if testCol == "BT" :
        #     fillYorN= True if testValue[0:1] not in ["1","4"] else False

        # #配件墙陈列
        # if testCol =="BU" and imageType=="2.":
        #     fillYorN= True if testValue[0:1] not in ["1","3"] else False
        # if testCol =="BV" and imageType=="1.":
        #     fillYorN= True if testValue[0:1] not in ["1","4"] else False         
        
        #人员
        if testCol == "BW" and testValue[0:1] != "1" and datiren=="1":
            fillYorN = True 

        #知识考核
        if datiren=="1":
            if testCol =="CA" and otherValue(testCol,rowNumber,-32)=="Y":#BCD
                if testValue.find("A、") != -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True
            if testCol =="CB":
                if otherValue(testCol,rowNumber,-2)=="Y" or otherValue(testCol,rowNumber,-33)=="Y":#B
                    fillYorN = True if testValue[0:1]!="B" else False

            if testCol =="CC":#
                fillYorN = True if testValue[0:1]!="D" else False

            if testCol =="CD" :#ABCD
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                        fillYorN = True
            if testCol =="CE":#
                fillYorN = True if testValue[0:1]!="A" else False
            if testCol =="CF":#
                fillYorN = True if testValue[0:1]!="C" else False    
            if testCol =="CG":#
                fillYorN = True if testValue[0:1]!="B" else False    
            if testCol =="CH":#
                fillYorN = True if testValue[0:1]!="D" else False    
            if testCol =="CI":#
                fillYorN = True if testValue[0:1]!="C" else False    
            if testCol =="CJ" :#ABCD
                if testValue.find("A、") == -1 or testValue.find("B、") != -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                        fillYorN = True

        #竞品 CK 1
        
        # 违规  
        if testCol == "CM" and testValue[0:2]not in ["1.", "2.", "3."]:
            fillYorN = True
        if testCol == "CO" and testValue[0:2]not in ["7.", "8.", "9."]:
            fillYorN = True
        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            queDic["C{}".format(rowNumber)].append(
                "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
for r in range(2, ws2.max_row+1):
    valueIn = " ".join(queDic["C{}".format(r)])
    ws2["C{}".format(r)].value = valueIn

# 保存
fileNameNew = "check.xlsx"
print("saving......")
wb.save(fileNameNew)

wb.close()
