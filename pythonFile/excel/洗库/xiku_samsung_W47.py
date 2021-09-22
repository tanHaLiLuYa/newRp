import datetime
import os
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from tqdm import tqdm


def listInlist(list1, bigList):
    a = []
    for i in list1:
        listMa = re.findall(r"[\u4e00-\u9fa5]{2,8}|[0-9]{3,4}X{0,1}", i)
        if len(listMa) == 2:
            # 调整顺序 不怕乱序
            nameId = listMa[0] + \
                listMa[1] if listMa[1][0].isdigit() else listMa[1] + listMa[0]
            if nameId not in bigList:
                a.append(i)
        else:
            a.append(i)
    return a


def N20YANGJI(list1, list2):
    if len(set(list1)) != len(list1) or len(list2) % 2 != 0:
        return True
    else:
        listTest = []
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])

        for i in listTest:
            if i[0].isdigit():
                if int(i[0]) > 2 or (int(i[0]) > 1 and i[1] != "金"):
                    return True
            else:
                if int(i[1]) > 2 or (int(i[1]) > 1 and i[0] != "金"):
                    return True
        return False


def GESHIHUA(list2):
    listTest = []
    if len(list2) > 2:
        for i in range(0, len(list2), 2):
            listTest.append(list2[i]+list2[i+1])
        return ",".join(listTest)
    else:
        return ",".join(list2)

        # listTest = []
        # for i in range(0,len(list2),2):
        #     listTest.append(list2[i]+list2[i+1])
        # for i in listTest:
        #     if i[0].isdigit():


def strToDicForTab(strIn):
    a = {}
    a["T720"] = strIn[strIn.find("T720")+4] if strIn.find("T720") != -1 else 0
    a["T510"] = strIn[strIn.find("T510")+4] if strIn.find("T510") != -1 else 0
    a["T860"] = strIn[strIn.find("T860")+4] if strIn.find("T860") != -1 else 0
    a["P610"] = strIn[strIn.find("P610")+4] if strIn.find("P610") != -1 else 0

    return a


def testTabFun(dic, dic_):
    for i in ["P610", "T860", "T510", "T720"]:
        if int(dic[i]) != int(dic_[i]):
            return True
    return False


def otherValue(letter,r, num):
    tNum = column_index_from_string(letter)
    letternew= get_column_letter(tNum+num)
    return ws2["{}{}".format(letternew,r)].value


print("loading......")
fillStyle = PatternFill("solid", fgColor="E6B8B7")

root_path = r"D:\work\samsung\W48周报\洗库"

os.chdir(root_path)

fileName = "W48洗库_1124_1450.xlsx"

# matchPatternTab = r"P610{ }[0-9]{1}"
matchPattern = r"[A-W]{1,2}[1-9]{1,3}.{0,1}[1-9]{0,1}[.]"
matchPatternID = r"[\u4e00-\u9fa5]{2,10}.{0,4}[0-9]{3,4}[X]{0,1}"

wb = load_workbook(fileName, data_only=True)
# print(wb.sheetnames)

ws2 = wb.copy_worksheet(wb["提交"])
ws2.title = "check1"


queDic = {}
titleDic = {}

# 开始列 X,结束列BL,问题记录成字典下的列表
starColumnName = "T"
endColumnName = "CG"  # 不能取空列

# 读取表头第一行 记录题目的 字母+数字 标识 为字典
for c in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
    if re.search(matchPattern, ws2["{}1".format(get_column_letter(c))].value):
        titleDic["{}".format(get_column_letter(c))] = re.search(
            matchPattern, ws2["{}1".format(get_column_letter(c))].value).group().replace(".", "")
# print(titleDic)
# '''

# 开始主程序          ��
fillYorN = False
for rowNumber in range(2, ws2.max_row+1):
    queDic["C{}".format(rowNumber)] = []
    for colNumber in range(column_index_from_string(starColumnName), column_index_from_string(endColumnName)+1):
        # ————————————————————————————————————————————————————————————————————————————
        testCol = get_column_letter(colNumber)
        print("列：", testCol, "行：", rowNumber)
        testValue = str(ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value) \
            if ws2["{}{}".format(get_column_letter(colNumber), rowNumber)].value != None else "00"

        investorType = ws2["K{}".format(rowNumber)].value
        investorType = investorType[0:3]  # 投资类型
        imageType = ws2["T{}".format(rowNumber)].value
        imageType = imageType[0:2]  # 新旧形象
        datiren = ws2["BQ{}".format(rowNumber)].value if ws2["BQ{}".format(rowNumber)].value else "none"
        datiren =datiren[0:1]
        # ————————————————————————————————————————————————————————————————————————————

        # 题目列
        # —————————————————————————————————————————————————————————————————————————————————————————————————————
        # 身份证
        if testCol == "X":
            testIDList = otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            if not re.match(matchPatternID,  testValue):
                fillYorN = True
            elif listInlist(re.findall(matchPatternID, testValue), testIDList):
                ws2["{}{}".format(get_column_letter(
                    colNumber), rowNumber)].fill = fillStyle
                queDic["C{}".format(rowNumber)].append("请检查{}列{}题，{}未匹配；".
                                                       format(testCol, titleDic[testCol], listInlist(re.findall(matchPatternID, testValue), testIDList)))

        # 1 门店形象 店外露出
        # ————————————————————————————————————————————————————————————————————————————

        if testCol in ["AB", "Z", "AA","AC","AV","BN","BO","BZ"] and testValue[0:1] != "1" :  #
            fillYorN = True
        # if testCol in ["U"] and testValue[0:2] not in ["1.", "5."]:
        #     fillYorN = True
        if testCol in ["AD"] and testValue[0:1] not in ["1", "4"] :
            fillYorN = True

        # 体验台销售柜
        if testCol in["AG", "AJ", "AM", "AP"] :
            touru = otherValue(testCol,rowNumber, -1) if otherValue(testCol,rowNumber, -1) else 0
            fillYorN = True if int(touru) != int(testValue) else False
        # S20 FE A4 立牌 AV
        # S20 FE KV 
        if testCol == "AW" and testValue[0:1] not in ["1", "3"] and investorType == "SIS" :
            fillYorN = True
        # Fold2 机模/真机
        if testCol == "AX" and otherValue(testCol,rowNumber,-4)=="Y":
            fillYorN = True if testValue[0:1] != "1" else False
        # S20 FE 机模/真机
        if testCol == "BA" and otherValue(testCol,rowNumber,-1)=="Y":
            fillYorN = True if testValue[0:1] != "1" else False
        # S20 FE 多色
        if testCol== "BB" and imageType=="2." and otherValue(testCol,rowNumber,-2)=="Y":
            tiyantai = int(otherValue(testCol,rowNumber,2))
            if investorType=="SIS" and tiyantai>=3:
                fillYorN= True if testValue[0:1]!="1" else False
            else:
                fillYorN= True if testValue[0:1] not in ["1","2"] else False
        if testCol =="BC"  and imageType=="1." and otherValue(testCol,rowNumber,-3)=="Y":
            if investorType=="SIS" :
                fillYorN= True if testValue[0:1]!="1" else False
            else:
                fillYorN= True if testValue[0:1] not in ["1","2"] else False            
        #新品台
        if testCol =="BE" and investorType=="SIS" and imageType=="1." and otherValue(testCol,rowNumber,-11)=="Y":
            fillYorN= True if testValue[0:1] not in ["1","4"] else False    
        #4.8 S20 FE Family集中陈列
        if testCol =="BF" and investorType=="SIS" and imageType=="2.":
            fillYorN= True if testValue[0:1] not in ["1","4","3"] else False    
        #4.9 传统形象门店S20 FE 陈列套装
        if testCol =="BG"  and imageType=="2.":
            fillYorN= True if testValue[0:1]!="1" else False
        #4.10 新形象门店S20 FE 3机托架陈列
        if testCol =="BH"  and imageType=="1.":
            fillYorN= True if testValue[0:1] not in ["1","6"] else False   
            if testValue[0:1]=="6" and investorType=="SIS":
                fillYorN= True
        #4.14 体验台桌下灯箱与产品保持同侧
        if testCol =="BI" and investorType=="SIS" and imageType=="2.":
            fillYorN= True if testValue[0:1] not in ["1","5","3"] else False    
        #retail
        if testCol == "BJ" :
            fillYorN= True if testValue[0:1] not in ["1","4"] else False
        if testCol == "BK" :
            fillYorN= True if testValue[0:1] not in ["1","5"] else False
        if testCol == "BL" :
            fillYorN= True if testValue[0:1] not in ["1","6"] else False
        
        #a机模陈列
        if testCol == "BM" :
            fillYorN= True if testValue[0:1] not in ["1","4"] else False
            if len(testValue)>11:
                fillYorN= True 
        #   4.16 产品陈列规范 BN	1
        #   4.17 机位空置 BO	1
        #人员
        if testCol == "BP" and testValue[0:1] != "1":
            jiedairen= otherValue(testCol,rowNumber,1) if otherValue(testCol,rowNumber,1) else "none"
            fillYorN = True if jiedairen[0:1] == "1" else False
        #知识考核
        if datiren=="1":
            if testCol =="BS":#ACD
                if testValue.find("A、") == -1 or testValue.find("B、") != -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True
            if testCol =="BT":#ABCD
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True
            if testCol in ["BV","BX","BY"]:#ABC
                if testValue.find("A、") == -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") != -1:
                    fillYorN = True
            if testCol =="BW":#BCD
                if testValue.find("A、") != -1 or testValue.find("B、") == -1 or testValue.find("C、") == -1 \
                        or testValue.find("D、") == -1:
                    fillYorN = True
            if testCol =="BU":#A
                fillYorN = True if testValue[0:1]!="A" else False

        # 违规
        #竞品 BZ
        if testCol == "CB" and testValue[0:2]not in ["1.", "2.", "3."]:
            fillYorN = True
        if testCol == "CD" and testValue[0:2]not in ["7.", "8.", "9."]:
            fillYorN = True
        # 填充颜色 记录问题到B列
        # ————————————————————————————————————————————————————————————————————————————
        if fillYorN:
            queDic["C{}".format(rowNumber)].append(
                "请检查{}列{}题；".format(testCol, titleDic[testCol]))
            ws2["{}{}".format(get_column_letter(colNumber),
                              rowNumber)].fill = fillStyle
            fillYorN = False

# print(queDic)
# 写入数据
for r in tqdm(range(2, ws2.max_row+1)):
    valueIn = " ".join(queDic["C{}".format(r)])
    ws2["C{}".format(r)].value = valueIn

# 保存
fileNameNew = "check.xlsx"
print("saving......")
wb.save(fileNameNew)

wb.close()
