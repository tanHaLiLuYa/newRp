#!/usr/bin/env python
# encoding: utf-8
from openpyxl import load_workbook
import os
import re


def cleanOption(allOption):
    needOpt = re.compile("[ABCDEFG]、")
    optlist = needOpt.findall(str(allOption))
    mystr = ''.join(sorted(optlist))
    # return mystr  #这是返回带、的
    return mystr.replace("、", "")


# excelpath = "/Users/simon/_work/CBIC/custom/ThinkPad/FY18Q2_ThinkPad_店面规范检查_第6次知识考核.xlsx"
excelpath = "/Users/simon/_work/CBIC/1-custom/hp/toABCD/FY19Q3_HP_2nd_知识考核-汇总.xlsx"

print(os.getcwd())
# print(cleanOption("B、 X系列,C、E系列 ,A、 X1系列  ,D、 S系列,"))

wb = load_workbook(filename=excelpath)
ws = wb.get_sheet_by_name('检查明细')
print(wb.get_sheet_names())

areas = ws["Q4":"DL974"]
for row in areas:
    for cell in row:
        old_value = cell.value
        cell.value = cleanOption(old_value)
        print("--已经修改为{}".format(cell.value))
    print("===========当前行修改完毕")
print("=====================所有单元格修改完毕")

print("另保存为新文件......")
# wb.save('/Users/simon/_work/CBIC/custom/ThinkPad/FY18Q2_ThinkPad_店面规范检查_第6次知识考核_ABCD.xlsx')
wb.save('/Users/simon/_work/CBIC/1-custom/hp/toABCD/FY19Q3_HP_2nd_知识考核-汇总_ABCD.xlsx')
print("新文件保存成功")
