#!/usr/bin/env python
# encoding: utf-8


"""
@version: ??
@author: Fenghua Ye
@license: Apache Licence
@contact: wildplant@gmail.com
@site: http://www.fangworks.com
@software: PyCharm
@time: 2018-11-7 00:30

从一个表格里面找到照片的链接
下载到本地
命名为 省--市--店面ID-中文店名/姓名.照片后缀

"""
import os
import shutil
import sys
import requests
from openpyxl import load_workbook

save_root = r"G:\三星\MCS_PICS\导购员20190527"
# save_root = r"/Users/simon/Desktop/sam-photo/"
small_pic = r"G:\三星\MCS_PICS\导购员20190527_小照片"
# 一个函数，把一个链接下载到本地的path


def get_status(link):
    try:
        headers = {'Connection': 'close', }
        request = requests.get(link, headers=headers)
        status = request.status_code
        return status
    except:
        return 0


def down_save(link, path):
    try:
        headers = {'Connection': 'close', }
        request = requests.get(link, headers=headers)
        img = request.content
        with open(path, "wb") as f:
            f.write(img)
    except Exception as e:
        print("出错了---", e)


# down_save("https://www.samsungmcs.com/mcs/photos/2018/01/15/20180115121502999_0.jpg", "./1.jpg")

"""
有些错误可能是因为店名或部分路径中存在空格，引致无法找到路径
"""

# 读取excel表
excelpath = "G:\三星\MCS_PICS\照片库20190528.xlsx"
wb = load_workbook(filename=excelpath)
ws = wb.get_sheet_by_name('photo')


def get_people_pic(start, end):
    for i in range(start, end):
        link_value = ws[f"C{i}"].value
        if len(str(link_value)) < 5:
            pass
        else:
            photo_folder = ws[f"P{i}"].value + "__" + ws[f"Q{i}"].value
            if not os.path.exists(os.path.join(save_root, photo_folder)):
                os.makedirs(os.path.join(save_root, photo_folder))
            file_name = ws[f"B{i}"].value + "__" + \
                ws[f"A{i}"].value + "." + link_value.split(".")[-1]
            file_path = os.path.join(save_root,  photo_folder, file_name)
            if os.path.exists(file_path) and os.path.getatime(file_path) > 0:
                ws[f"CG{i}"].value = 1
                pass
                try:
                    fsize = os.path.getsize(file_path)
                    if fsize > 0:
                        ws[f"CH{i}"].value = round(fsize/1024, 2)
                        # print(f"第{i}行图片大小为 {fsize/1024}K")
                        pass
                except Exception as e:
                    print(e)
            elif ws[f"CG{i}"].value == 1:
                # print(f"----第{i}---已经下载过")
                pass
            elif get_status(link_value) == 200:
                # ws["Y{}".format(str(i))].value = get_status(link_value)

                down_save(link_value, file_path)
                try:
                    fsize = os.path.getsize(file_path)
                    if fsize > 0:
                        ws[f"CG{i}"].value = 1
                        ws[f"CH{i}"].value = round(fsize/1024, 2)
                        print(f"第{i}行的图片{file_path}--已经保存")
                except Exception as e:
                    print(e)

            else:
                ws[f"CG{i}"].value = 0
    wb.save(excelpath)
    print("文件已经保存")


def get_store_pic(start, end):
    for i in range(start, end):
        store_link = ws[f"CG{i}"].value
        # print(store_link_list)

        if store_link is None or store_link == "":
            pass
        elif ws[f"CI{i}"].value == 1:
            pass
        elif ws[f"CI{i}"].value != 1:
            store_link_list = store_link.split(",")
            for store_link in store_link_list:
                if get_status(store_link) != 200:
                    pass
                else:
                    photo_folder = ws[f"O{i}"].value + "__" + ws[f"P{i}"].value
                    if not os.path.exists(os.path.join(save_root, photo_folder)):
                        os.makedirs(os.path.join(save_root,  photo_folder))
                    file_name = "店面照片" + "__" + \
                        ws[f"P{i}"].value + "__" + store_link.split("/")[-1]
                    file_path = os.path.join(
                        save_root, photo_folder, file_name)

                    down_save(store_link, file_path)
                    print(file_path)
                    try:
                        if os.path.getsize(file_path) > 0:
                            ws[f"CI{i}"].value = 1
                        else:
                            ws[f"CI{i}"].value = 0
                    except Exception as e:
                        print(e)
                    print(f"第{i}行的图片{file_path}--已经保存")

    wb.save(excelpath)
    print("文件已经保存")


def del_empty_folder(path):
    files = os.listdir(path)
    for each_f in files:
        f = os.path.join(path, each_f)
        if "Thumbs.db" in str(f):
                os.remove(f)
                print(f, "Thumbs.db文件删除")
    for each_f in files:
        f = os.path.join(path, each_f)
        if os.path.isdir(f):
            if not os.listdir(f):
                print(f"正在准备删除----{f}")
                os.rmdir(f)

def copy_small(size_k, path):
    for root, dirs, files in os.walk(path):
        for f in files:
            fname = os.path.join(root,f)
            if os.path.getsize(fname)<=int(size_k)*1024:
                desc_root = root.replace(save_root,small_pic)
                desc_fname = os.path.join(desc_root,f)
                if not os.path.exists(desc_root):
                    print("正在创建目录", desc_root)
                    os.makedirs(desc_root)
                print(f, root, desc_root)
                shutil.copy(fname, desc_fname)
            else:
                pass



class Main():
    def __init__(self):
        pass


if __name__ == '__main__':
    # get_store_pic(2,2707)
    # start_row = int(sys.argv[1])
    # end_row = int(sys.argv[2])
    # get_people_pic(start_row,end_row)
    # for row in range(35, 76):
    # get_people_pic(row*100, (row+1)*100)
    del_empty_folder(save_root)
    # get_people_pic(2,7633)

    # copy_small(100,save_root)
