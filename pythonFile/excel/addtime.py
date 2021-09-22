import datetime
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

def plustime(col,strin):
    v = str(strin)
    if len(v)>8:
        t = datetime.datetime.strptime(v[-8:], "%H:%M:%S")
    else:
        t = datetime.datetime.strptime(v, "%H:%M:%S")
    if col % 2 == 0:#偶数列是开始时间 加15分钟
        return (t + datetime.timedelta(minutes=15)).strftime("%H:%M")
    else:#奇数列是结束时间 减15分钟
        return (t + datetime.timedelta(minutes=-15)).strftime("%H:%M")

def doubleplustime(strin):
    strin = strin.replace("~","-")
    strin = strin.replace("～","-")
    a = strin.split("-")
    if len(a)==2:
        if len(a[0]) > 5:
            t = datetime.datetime.strptime(a[0], "%H:%M:%S")
        else:
            t = datetime.datetime.strptime(a[0], "%H:%M")
        a[0]=(t + datetime.timedelta(minutes=15)).strftime("%H:%M")
        
        if len(a[1]) > 5:
            t = datetime.datetime.strptime(a[1], "%H:%M:%S")
        else:
            t = datetime.datetime.strptime(a[1], "%H:%M")
        a[1]=(t + datetime.timedelta(minutes=-15)).strftime("%H:%M")

        return "-".join(a)
    else:
        print("error")
        global errorfind
        errorfind = True
        return "error " + strin


chinesechar = re.compile(r"[\u4e00-\u9fa5]{1,6}") 
root_path = r"E:\work\tpp\samsung\2021年\汇总\2020年数据-至今\执行汇总总结"

os.chdir(root_path)

fileName = "W29导购员排班表_0712已对比.xlsx"
# 数据格式 调整为 18:00时间格式

wb = load_workbook(fileName)
ws = wb.active

quitRe =[]

# input
begincol = "H"
beginnum =column_index_from_string(begincol)
endcol   = "Q"
endnum   =column_index_from_string(endcol)

#开始列是否为偶数
startWTColIsEven = True if beginnum % 2 == 0 else False

errorfind = False
#主循环体
for r in range(3,ws.max_row+1):#第三行开始
    for c in range(beginnum,endnum+1):#
        ce = ws.cell(r,c)
        print(r,c)
        if ce.data_type == "s":
            if re.search(chinesechar, str(ce.value)):
                # print("休息请假类：",ce.value)
                quitRe.append(ce.value)
            else:
                # print("两段时间：",ce.value)
                ce.value =  doubleplustime(str(ce.value))
        elif ce.data_type == "d":
            # print("标准时间：",ce.value)
            ce.value = plustime(c,ce.value) if startWTColIsEven else plustime(c+1,ce.value)
        else:
            print("error is finding")
            errorfind = True
            break
        
quitRe = list(set(quitRe))

print(quitRe,"\n",errorfind)

wb.save(fileName.replace(".xlsx","-修改.xlsx"))

wb.close()