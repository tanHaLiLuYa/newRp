#!/usr/bin/env python
# encoding: utf-8
from openpyxl import load_workbook
import os
import re
import shutil

'''
实现功能：
文件夹下是很多店面，要把店面按ID号归类到大区里；ID号和大区的对应关系在excel里；


思路：
1、遍历文件夹，找出文件夹的名称（前8位）
2、在excel中读取包含ID和大区的区域，组成一个词典{}；
3、如果文件夹的名称在词典里，进行拷贝
4、新文件夹的路径为放置路径+词典中的值

'''

excel_path = "/Users/simon/Documents/12月intel/ours.xlsx"
folder_path = "/Users/simon/Documents/12月intel/12月消费自检照片intel已挑选/"

# 读取excel组成一个词典
wb = load_workbook(filename=excel_path)
ws = wb.get_sheet_by_name('ours')
print(wb.get_sheet_names())
cells = ws["B2":"D1785"]
our_dict = {}
for row in cells:
    id_area_row = list(x.value for x in row)
    # print(id_area_row)
    our_dict[id_area_row[2]] = id_area_row[0]
print(our_dict)
# for cell in row:
#     our_list.append(cell.value)



# 读取文件夹

list_dirs = os.walk(folder_path)

for root, folders, files in list_dirs:
    for folder in folders:
        store_id = folder[:8]
        area = our_dict[store_id]
        id_path = os.path.join(root, folder)
        new_path = os.path.join(folder_path, area)
        if area:
            print(store_id, id_path, "店面的区域为", area, "，将要放到", new_path)
            if not os.path.exists(new_path):
                os.makedirs(new_path, exist_ok=True)
                print("创建新文件夹成功")
            else:
                shutil.move(id_path, new_path)
                print("已经移动到新文件夹中")
        else:
            print("有可能店面ID是错的")
