#!/usr/bin/env python
# encoding: utf-8
import os
import sys
import datetime
import time

from openpyxl import load_workbook

people_sheet = "/Users/simon/_work/CBIC/custom/samsung/201904/人员异常对照的python研究/W19排班表.xlsx"
ab_sheet = "/Users/simon/_work/CBIC/custom/samsung/201904/人员异常对照的python研究/w19检查人数1447.xlsx"

# 各个日期在排班表中所对应的列
day_dict = {
    "6": {"start": "M", "end": "N"},
    "7": {"start": "Q", "end": "R"},
    "8": {"start": "U", "end": "V"},
    "9": {"start": "Y", "end": "Z"},
    "10": {"start": "AC", "end": "AD"},
    "11": {"start": "AG", "end": "AH"},


}
print(day_dict["9"]["start"])
wb_a = load_workbook(filename=ab_sheet)

ws_a = wb_a['data']
ws_people = load_workbook(filename=people_sheet)['data']


for i in range(2, ws_a.max_row + 1):
    ap_id = ws_a[f"E{i}"].value
    ap_date = ws_a[f"B{i}"].value
    ap_datetime = datetime.datetime.combine(ap_date,ws_a[f"C{i}"].value)
    # print(str(i),ap_date.year, ap_date.month, ap_date.day)

    for j in range(3, ws_people.max_row + 1):
        p_id = ws_people[f'D{j}'].value
        if ap_id == p_id:
            time_col_start = day_dict[str(ap_date.day)]["start"]
            time_col_end = day_dict[str(ap_date.day)]["end"]
            time_start_cell = ws_people[time_col_start + f"{j}"].value
            time_end_cell = ws_people[time_col_end + f"{j}"].value
            if time_start_cell == "休假":
                ws_a[f"W{i}"].value = "休假"
                ws_a[f"X{i}"].value = "休假"
                ws_a[f"Y{i}"].value = "休假中"
            else:             
                time_start_str = f"{ap_date.year}-{ap_date.month}-{ap_date.day} {time_start_cell}"
                time_end_str = f"{ap_date.year}-{ap_date.month}-{ap_date.day} {time_end_cell}"
                time_start_value = datetime.datetime.strptime(time_start_str, "%Y-%m-%d %H:%M")
                time_end_value = datetime.datetime.strptime(time_end_str, "%Y-%m-%d %H:%M")
                if ap_datetime > time_start_value and ap_datetime < time_end_value:
                    print(f"{i}",ap_datetime, time_start_value, time_end_value, "----", True)
                    ws_a[f"W{i}"].value = time_start_value
                    ws_a[f"X{i}"].value = time_end_value
                    ws_a[f"Y{i}"].value = "导购员在上班"
                else:
                    print(f"{i}", ap_datetime, time_start_value,
                          time_end_value, "<<<<<<", False)
                    ws_a[f"W{i}"].value = time_start_value
                    ws_a[f"X{i}"].value = time_end_value
                    ws_a[f"Y{i}"].value = "访问时间不在上班时间内"
wb_a.save(ab_sheet)
