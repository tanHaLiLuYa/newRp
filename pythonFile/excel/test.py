# encoding:utf-8

import os
import re
import sys
import csv
from tqdm import tqdm 
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Color, Font, colors
from openpyxl.utils import column_index_from_string, get_column_letter
import xlwings as xw
'''
file_path = r"E:\work\samsung\2018-W46-\19_1209波哥\导出原库"
file_dirs = os.listdir(file_path)

app = xw.App(visible=False, add_book=False)

# 正则表达式匹配店面ID
PatternId = r"CS\d{7}"


def isID(CellValue):
    likeID = re.compile(PatternId)
    result = likeID.search(CellValue)
    if result:
        return True
    else:
        return False


def getcloumn(listIn, strIn):
    for i in range(len(listIn)):
        if listIn[i]:
            if strIn in listIn[i]:
                return i


for file in file_dirs:
    wb = app.books.open(os.path.join(file_path, file))
    ws = xw.sheets.active
    List_head = ws.range("A1:T1").value
    
    DoFile=file[:7] if file[:2] =="20" else  (str(2019)+file[:3])
    InterviwerCol = getcloumn(List_head, "访问员姓名和电话")
    
    rng =ws.range("A1").expand("table")
    max_row =rng.rows.count
    for i in tqdm(range(2,max_row+1)):
        List_second = ws.range(f"A{i}:G{i}").value

        StoreIdCol = getcloumn(List_second, "CS0")

        StroeIdValue = ws.range(i,StoreIdCol+1).value
        InterviwerValue = ws.range(i,InterviwerCol+1).value

        if InterviwerValue:
            with open(os.path.join(r"E:\work\samsung\2018-W46-\19_1209波哥\新建文件夹","myfile.txt"),"a",newline='',encoding="utf-8") as f:
                row = [DoFile,StroeIdValue,InterviwerValue]
                for i in row :
                    f.write(i+"!")
                f.write("\n")
        
    print(f"{max_row}行数据已处理好",file)

    wb.save()
    wb.close()
app.quit()
# Aimlist = []

        # print(os.path.join(root,file))
# for file in file_dirs:
#     wb = openpyxl.load_workbook(os.path.join(file_path,file),read_only=True)
#     wsname =wb.sheetnames[0]
#     ws         =wb[wsname]
#     head_area=ws["A1:Z1"]
#     second_area =ws["A2:Z2"]
#     print(file,"已经打开了")
#     # ColFangWen = 0
#     for i in range(0,26):
#         # print(head_area[0][i].value)
#         if "访问员姓名和电话" in head_area[0][i].value:
#             ColFangWen=i

#             print(i,"是访问员列的列数+++++",file)
#         if  second_area[0][i].value:
#             if "CS" in second_area[0][i].value:
#                 IDcol = i
#                 print(i,"是商场ID列数")
#     # print(ws.max_row)
#     print("正在处理",file)
#     DoFile=file[:7] if file[:2] =="20" else  (str(2019)+file[:3])
#     for r in tqdm(range(2,ws.max_row+1)):
#         # for c in range(1,ColFangWen+1):
#             # print("++++++")
#         CellValue = str(ws.cell(row =r, column = ColFangWen + 1).value)
#         IDValue =str(ws.cell(row =r, column = IDcol + 1).value)
#         if CellValue:
#             with open(os.path.join(r"E:\work\samsung\2018-W46-\19_1209波哥\新建文件夹","myfile.csv"),"a",newline='') as f:
#                 row = [DoFile,IDValue,CellValue]
#                 # print(row)
#                 write=csv.writer(f)
#                 write.writerow(row)

#     print("_____+++++处理成功",file)
#     wb.close()


'''