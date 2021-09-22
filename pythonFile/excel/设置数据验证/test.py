import datetime
import os
import re
import sys
import time
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

path=r"D:\work\samsung\W47周报\洗库\合并\output\新建 Microsoft Excel 工作表.xlsx"
wb = load_workbook(path)
ws =wb.active
dav1=openpyxl.worksheet.datavalidation.DataValidation(type = 'list',
                                                             formula1 = '"12,9,0,NA"',allow_blank=True)  
                                                                                                           
dav2=openpyxl.worksheet.datavalidation.DataValidation(type = 'list',
                                                             formula1 = '"4,6,0,NA"',allow_blank=True)  
# dav.error ='请从下拉列表中选择'
# dav.errorTitle = '错误输入'

# Optionally set a custom prompt message
                                                                                                                                    

ws.add_data_validation(dav1)
ws.add_data_validation(dav2)

dav1.add("A9")
dav1.add("B9:C10")
dav2.add("E9:E10")
c1 = ws["A9"]
c1.value = "Dog"
path=r"D:\work\samsung\W47周报\洗库\合并\output\新建 Microsoft Excel 工作表-1.xlsx"
wb.save(path)