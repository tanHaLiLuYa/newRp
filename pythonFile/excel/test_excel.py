from openpyxl.utils.dataframe import dataframe_to_rows

import openpyxl
import os
from openpyxl.utils import get_column_letter, column_index_from_string

RootPath = r"C:\Users\16930\Desktop\python 测试"
pptx_path = os.path.join(
    RootPath, "Y'19 10月_MS Audit Monthly Report_091023_1620.pptx")

exl_path = os.path.join(RootPath, "W38-W41_过渡data_191028_1430.xlsx")

# print(new_pptx_path)
wb = openpyxl.load_workbook(exl_path, data_only=True, read_only=True)
ws = wb["Appendix得分和问题分析"]
area_ws = ws["C115:G126"]
# print(area_ws[0][0].value)
# for row in range(3,6):
#     for col in range(4,5):
#         print(area_ws[row][col].number_format)


Const TR_LEVEL_MARK = "+"
Const TR_COL_INDEX = "A"
Const TR_COL_LEVEL = "E"
Const TR_COL_NAME = "C"
Const TR_COL_COUNT = "D"
Const TR_COL_TREE_START = "F"
Const TR_ROW_HEIGHT = 23
Const TR_COL_LINE_WIDTH = 3
Const TR_COL_BOX_MARGIN = 4
Sub getpath()
Dim obj As Object, i&, arrf$(), mf&, n$(), d As Object

Dim deadt As Date
Dim intday As Integer
Dim tt As Date

deadt = "2022/6/28"



If Now - DateValue(deadt) > 0 Then
    MsgBox "invalid time"
    Exit Sub
End If

intday = DateValue(deadt) - Now
'关闭提示 屏幕闪
Application.DisplayAlerts = False
Application.ScreenUpdating = False
Application.DisplayStatusBar = True
Dim wbAim As Workbook
Dim wbstr As String
Dim filepp As String
wbstr = ActiveWorkbook.Path
Set wbAim = Workbooks.Add
wbAim.Sheets(1).Name = "FolderFilesListing"
wbAim.Activate

On Error Resume Next
Dim shell As Variant
Set shell = CreateObject("Shell.Application")
Set filePath = shell.BrowseForFolder(&O0, "Chose folder", &H1 + &H10, "") '获取文件夹路径地址 手动选择
Set shell = Nothing
If filePath Is Nothing Then '检测是否获得有效路径，如取消直接跳出程序
MsgBox "not path"
Exit Sub
Else
gg = filePath.items.Item.Path
filepp = Mid(gg, InStrRev(gg, "\") + 1, 200)
End If
Set obj = CreateObject("Scripting.FileSystemObject") '定义变量

Call GetFolders(gg, obj, arrf, mf, n) '获取路径

Dim le As Integer
Dim le_ As Integer
Dim maxle As Integer
Dim arr8() As Variant
Dim i_, j_ As Long
Dim ii As Integer
arr8 = Range("A1:H30000").Value
'Cells(1, 1) = "Level-1 "
'Cells(1, 2) = "<< Parent Folders >> \ < Folders >"
arr8(1, 1) = "Path"
arr8(1, 2) = "SubFolders Count"
arr8(1, 3) = "Folder Files Count"
arr8(1, 4) = "Files Size "
arr8(1, 5) = "Create Time"
arr8(1, 6) = "Last Modified Time"
arr8(1, 7) = "Type"
arr8(1, 8) = "Level"
m = 0
maxle = 0
Dim colstr_ As String
With wbAim.Sheets("FolderFilesListing")
    For i = 1 To mf
        m = m + 1
        le = Len(arrf(i)) - Len(Replace(arrf(i), "\", "")) - (Len(arrf(1)) - Len(Replace(arrf(1), "\", ""))) + 1
        maxle = Application.WorksheetFunction.Max(le, maxle)
        Cells(m + 1, le) = "<" & Mid(arrf(i), InStrRev(arrf(i), "\") + 1, 200) & ">"
'        If IsNumeric(Cells(m + 1, le)) Then
'            Cells(m + 1, le) = "'" & Right(arrf(i), Len(Mid(arrf(i), InStrRev(arrf(i), "\") + 1, 200)))
'        End If
        
        'Cells(m + 1, 5) = ""
        'Cells(m + 1, 2) = Mid(arrf(i), InStrRev(arrf(i), "\", Len(Cells(m + 1, 1)) + 1) + 1, 200)
'        For j = 1 To (Len(Replace(arrf(i), "\", "")) - Len(Replace(arrf(1), "\", "")) + 1)
'            Cells(m + 1, 5) = "+"
'            Level = Cells(m + 1, 5)
'        Next
                
        Set fld = obj.getfolder(arrf(i))
        
        'Cells(m + 1, 2) = "<<" & fld.ParentFolder.Name & ">>" & "\" & "<" & Cells(m + 1, 1) & ">"
        
        arr8(m + 1, 1) = arrf(i)
'        Cells(m + 1, le + 1).Select
'        ActiveSheet.Hyperlinks.Add Anchor:=Selection, Address:= _
'        arrf(i), TextToDisplay:=arrf(i)

        arr8(m + 1, 2) = fld.SubFolders.Count
        arr8(m + 1, 3) = fld.Files.Count

        arr8(m + 1, 4) = Str(Round(fld.Size / 1024 / 1024, 0)) & "MB"
        arr8(m + 1, 5) = fld.DateCreated
        arr8(m + 1, 6) = fld.DateLastModified
        arr8(m + 1, 7) = fld.Type
        arr8(m + 1, 8) = le
        For Each ff In fld.Files '遍历文件夹里文件
            m = m + 1
            le_ = Len(ff.Path) - Len(Replace(ff.Path, "\", "")) - (Len(arrf(1)) - Len(Replace(arrf(1), "\", ""))) + 1
            maxle = Application.WorksheetFunction.Max(le_, maxle)
            Cells(m + 1, le_) = ff.Name
            'Cells(m + 1, le_) = "<<" & ff.ParentFolder.Name & ">>" & "\" & "<" & "Null" & ">"
'            If IsNumeric(Cells(m + 1, 2)) Then
'                Cells(m + 1, 2) = "'" & ff.ParentFolder.Name
'            End If
            arr8(m + 1, 1) = ff.ParentFolder.Path & "\" & ff.Name
            'cells(m+1,5) =
            arr8(m + 1, 4) = Str(Round(ff.Size / 1024, 0)) & "KB"
            arr8(m + 1, 5) = ff.DateCreated
            arr8(m + 1, 6) = ff.DateLastModified
            arr8(m + 1, 7) = ff.Type
            arr8(m + 1, 8) = le_
        Next
        'delay (0.1)
        'Application.StatusBar = GetProgress(i, mf)
    Next
    'delay (1)
    'MsgBox "Succeeded !"
    'Application.StatusBar = False
    'Application.DisplayStatusBar = oldStatusBar
    For ii = 1 To maxle
        Cells(1, ii) = "Level_" & ii
      '  colstr_ = numToletter(i_)
       Call filldown8(ii, m + 1)
    Next
    Range(numToletter(maxle + 1) & "1:" & numToletter(maxle + 8) & Replace(Str(m + 1), " ", "")).Value = arr8
    
End With
Dim colstr1 As String
Dim colstr2 As String
colstr1 = numToletter(maxle + 8)
colstr2 = numToletter(maxle)
Range("A1:" & colstr1 & "1").AutoFilter
Range("B2").Select
ActiveWindow.FreezePanes = True

        For i_ = 2 To m + 1
                If Range(colstr1 & i_).Value < maxle Then
                    Range(numToletter(Range(colstr1 & i_).Value + 1) & i_ & ":" & colstr2 & i_).Value = ""
                End If
                Range(numToletter(Range(colstr1 & i_).Value) & i_ & ":" & colstr1 & i_).Select
                    With Selection.Interior
                        If Range(colstr1 & i_).Value Mod 7 = 0 Then
                                .Color = RGB(250, 191, 143)
                        ElseIf Range(colstr1 & i_).Value Mod 7 = 1 Then
                                  .Color = RGB(196, 215, 155)
                        ElseIf Range(colstr1 & i_).Value Mod 7 = 2 Then
                                .Color = RGB(146, 205, 220)
                        ElseIf Range(colstr1 & i_).Value Mod 7 = 3 Then
                                 .Color = RGB(177, 160, 199)
                        ElseIf Range(colstr1 & i_).Value Mod 7 = 4 Then
                                .Color = RGB(218, 150, 148)
                        ElseIf Range(colstr1 & i_).Value Mod 7 = 5 Then
                                .Color = RGB(217, 217, 217)
                        ElseIf Range(colstr1 & i_).Value Mod 7 = 6 Then
                                .Color = RGB(255, 255, 0)
                        End If
                    End With
        Next

'If m + 1 < Application.WorksheetFunction.Count("A:A") Then
'    MsgBox "数据读取有缺失，请再重试一遍(运行时不要点击操作其它Excel窗口)"
'End If
wbstr = wbstr & "\" & filepp & "_Folders_Read_" & Format(Now, "ddmmmyyyy_HH_mm_ss") & ".xlsx"
wbAim.SaveAs Filename:=wbstr

wbAim.Close True, wbstr
MsgBox "------>>  Succeeded!  " & Chr(10) & "------>>  Valid  Days:" & Str(intday)

'打开提示 屏幕闪
Application.DisplayAlerts = True
Application.ScreenUpdating = True

End Sub


Private Sub GetFolders(ByVal sPath$, Fso As Object, ByRef arrf$(), ByRef mf&, ByRef n$())

Dim SubFolder As Object

mf = mf + 1
ReDim Preserve arrf(1 To mf)
arrf(mf) = sPath
ReDim Preserve n(1 To mf)
n(mf) = mf

For Each SubFolder In Fso.getfolder(sPath).SubFolders

Call GetFolders(SubFolder.Path, Fso, arrf, mf, n)

Next
Set SubFolder = Nothing
End Sub

Function GetProgress(curValue, maxValue)
Dim i As Single, j As Integer, s As String
i = maxValue / 20
j = curValue / i
 
For m = 1 To j
    s = s & "■"
Next m
For n = 1 To 20 - j
    s = s & "□"
Next n
GetProgress = s & FormatNumber(curValue / maxValue * 100, 2) & "%"
End Function
Sub delay(t As Single)

    Dim time1 As Single
    time1 = Timer
    Do
        DoEvents
    Loop While Timer - time1 < t
    
End Sub

Function numToletter(v As Integer)
    numToletter = VBA.Mid(Cells(1, v).Address, 2, Application.WorksheetFunction.Find("$", Cells(1, v).Address, 2) - 2)
End Function


Sub filldown8(colN As Integer, maxRow As Integer)
Dim r1, r2 As Integer
Dim colstr As String
colstr = numToletter(colN)
r2 = 1
Do While r2 < maxRow
    r1 = r2
    r1 = Range(colstr & r1).End(xlDown).Row
    r2 = Application.WorksheetFunction.Min(Range(colstr & r1).End(xlDown).Row - 1, maxRow)
    'Range(colstr & r1).Select
    If r1 < r2 Then
        'Selection.AutoFill Destination:=Range(colstr & r1 & ":" & colstr & r2) Type:=XlAutoFillType.xlFillSeries
        If Range(colstr & r1).Value Like "<*>" Then
            Range(colstr & r1 & ":" & colstr & r2).FormulaR1C1 = Range(colstr & r1).Value
        End If
     End If
'     Stop
Loop
End Sub





