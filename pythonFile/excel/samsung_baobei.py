#!/usr/bin/env python
# encoding: utf-8
import os
import sys
import datetime
import time

from openpyxl import load_workbook,Workbook
root_path = r"E:\work\samsung\W36周报\事后修改排班表\离岗报备"
our_file = os.path.join(
    root_path, "【W36】导购员异常和门店异常情况_CBIC_0909_1451.xlsx")
our_back_file = our_file.replace(".xlsx","_backup.xlsx")
sx_D = os.path.join(root_path, "20190909_离岗报备记录导出-D.xlsx")
sx_A = os.path.join(root_path, "离岗信息-A.xlsx")
sx_F = os.path.join(root_path, "全国特殊事件数据-F.xlsx")
sx_P = os.path.join(root_path,"W36导购员事后排班表_20190909.xlsx")

id_our_col = "H"
name_our_col = "I"
date_our_col = "K"
time_in_col = "L"
time_out_col = "M"



print(our_file,"\n", sx_A, "\n", sx_D, "\n", sx_F)

#判断检查时间与报备时间是否重合
def time_overlap(start1, end1, start2, end2):
    if (start1 <= end2) and (start2 <= end1):
        return 1
    else:
        return 0
#备份函数
def backup(in_file=our_file, outfile=our_back_file):
    wb_our = load_workbook(our_file)
    # wb_our.remove(["异常导购员List"])
    wb_our.save(our_back_file)



# 离岗报备——D 记录导出表的函数：
def ligangbaobei(our_file=our_file, sxfile=sx_D):
    wb_our = load_workbook(our_file)
    ws_our = wb_our["异常导购员List"]
    wb_sx = load_workbook(sxfile)
    ws_sx = wb_sx.active
    ws_our["AE1"].value = "离岗报备_判断"
    ws_our["AF1"].value = "离岗报备_详情"

    for row in range(2, ws_our.max_row + 1):

        id_our = ws_our[f"{id_our_col}{row}"].value
        date_our = ws_our[f"{date_our_col}{row}"].value
        time_in = ws_our[f"{time_in_col}{row}"].value
        time_out = ws_our[f"{time_out_col}{row}"].value

        time_start=str(date_our)[:11] + str(time_in)[:5]
        our_start = datetime.datetime.strptime(f"{time_start}", "%Y-%m-%d %H:%M")

        time_end=str(date_our)[:11] + str(time_out)[:5]
        our_end = datetime.datetime.strptime(f"{time_end}", "%Y-%m-%d %H:%M")
        # our_end = datetime.datetime.combine(date_our, time_out)
        # print(">>>>>",our_start, our_end)
        # print(">>>>>",time_end)
        baobei = ""
        for r2 in range(3, ws_sx.max_row + 1):
            id_sx = ws_sx[f"A{r2}"].value

            if id_sx == id_our:
                
                date_2 = ws_sx[f"C{r2}"].value
                time_begin_str = ws_sx[f"D{r2}"].value if len(ws_sx[f"D{r2}"].value) >=5 else "08:00"
                time_end_str = ws_sx[f"E{r2}"].value if len(
                    ws_sx[f"E{r2}"].value) >= 5 else "23:59"
                time_begin = datetime.datetime.strptime(f"{date_2} {time_begin_str}", "%Y-%m-%d %H:%M")
                time_end = datetime.datetime.strptime(f"{date_2} {time_end_str}", "%Y-%m-%d %H:%M")
                print("匹配离岗报备文件成功",row,r2,)
                if time_overlap(our_start, our_end, time_begin, time_end) == 1:
                    
                    note = f"《离岗报备记录表》 中第{r2}行，报备离岗时间为{time_begin},返岗时间为{time_end}\n"
                    baobei = baobei + note
                    ws_our[f"AE{row}"].value = "有时间重合"
                    ws_our[f"AF{row}"].value = baobei
                    print("+++++", baobei)
    wb_our.save(our_file)
                
#对离岗信息文件匹配
def ligangxinxi(our_file=our_file, sxfile=sx_A):
    wb_our = load_workbook(our_file)
    ws_our = wb_our["异常导购员List"]
    wb_sx = load_workbook(sxfile)
    ws_sx = wb_sx.active
    ws_our["AG1"]="离岗信息_判断"
    ws_our["AH1"]="离岗信息_详情"
    for row in range(2, ws_our.max_row + 1):

        id_our = ws_our[f"{id_our_col}{row}"].value
        date_our = ws_our[f"{date_our_col}{row}"].value
        time_in = ws_our[f"{time_in_col}{row}"].value
        time_out = ws_our[f"{time_out_col}{row}"].value

        time_start=str(date_our)[:11] + str(time_in)[:5]
        our_start = datetime.datetime.strptime(f"{time_start}", "%Y-%m-%d %H:%M")

        time_end=str(date_our)[:11] + str(time_out)[:5]
        our_end = datetime.datetime.strptime(f"{time_end}", "%Y-%m-%d %H:%M")
        # print(">>>>>",our_start, our_end)
        baobei = ""
        for r2 in range(2, ws_sx.max_row + 1):
            id_sx = ws_sx[f"N{r2}"].value

            if id_sx == id_our:

                date_2 = ws_sx[f"O{r2}"].value
                time_begin_str = ws_sx[f"P{r2}"].value if len(
                    ws_sx[f"P{r2}"].value) >= 5 else f"{date_2} 09:00"
                time_end_str = ws_sx[f"Q{r2}"].value if len(
                    ws_sx[f"Q{r2}"].value) >= 5 else f"{date_2} 23:59"
                time_begin = datetime.datetime.strptime( time_begin_str, "%Y/%m/%d %H:%M")
                time_end = datetime.datetime.strptime( time_end_str, "%Y/%m/%d %H:%M")
                print("匹对离岗信息文件成功",row, r2,)
                if time_overlap(our_start, our_end, time_begin, time_end) == 1:

                    note = f"《离岗信息》中第{r2}行，报备离岗时间为{time_begin},返岗时间为{time_end}\n"
                    baobei = baobei + note
                    ws_our[f"AG{row}"].value = "有时间重合"
                    ws_our[f"AH{row}"].value = baobei
                    print("+++++", baobei)
    wb_our.save(our_file)

#特殊事件文件匹对
def teshushijian(our_file=our_file, sxfile=sx_F):
    wb_our = load_workbook(our_file)
    ws_our = wb_our["异常导购员List"]
    wb_sx = load_workbook(sxfile)
    ws_sx = wb_sx.active
    ws_our["AI1"]="特殊事件_判断"
    ws_our["AJ1"]="特殊事件_详情"

    for row in range(2, ws_our.max_row + 1):
        name_our = ws_our[f"{name_our_col}{row}"].value
        date_our = ws_our[f"{date_our_col}{row}"].value
        time_in = ws_our[f"{time_in_col}{row}"].value
        time_out = ws_our[f"{time_out_col}{row}"].value
        
        time_start=str(date_our)[:11] + str(time_in)[:5]
        our_start = datetime.datetime.strptime(f"{time_start}", "%Y-%m-%d %H:%M")

        time_end=str(date_our)[:11] + str(time_out)[:5]
        our_end = datetime.datetime.strptime(f"{time_end}", "%Y-%m-%d %H:%M")
        # print(">>>>>",our_start, our_end)
        baobei = ""
        for r2 in range(2, ws_sx.max_row + 1):
            name_sx = ws_sx[f"A{r2}"].value
            section = ws_sx[f"B{r2}"].value
            if name_sx == name_our:
                
                time_begin_str = ws_sx[f"C{r2}"].value
                time_end_str = ws_sx[f"D{r2}"].value
                time_begin = datetime.datetime.strptime(time_begin_str, "%Y-%m-%d %H:%M")
                time_end = datetime.datetime.strptime(time_end_str, "%Y-%m-%d %H:%M")
                print("匹对特殊事件成功",row, r2,)
                if time_overlap(our_start, our_end, time_begin, time_end) == 1:

                    note = f"《特殊事件数据》中第{r2}行，姓名{name_sx}，部门为{section}，报备离岗时间为{time_begin},返岗时间为{time_end}\n"
                    baobei = baobei + note
                    ws_our[f"AI{row}"].value = "有时间重合"
                    ws_our[f"AJ{row}"].value = baobei
                    print("+++++", baobei)
    wb_our.save(our_file)

# backup()
# ligangbaobei()
# ligangxinxi()
# teshushijian()
